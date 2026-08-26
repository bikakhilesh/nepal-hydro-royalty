#!/usr/bin/env python
"""Nepal hydropower ledger - scrape, reconcile, render.

    python hydro.py                 rebuild the dashboard from existing CSVs (default)
    python hydro.py build           same, explicitly
    python hydro.py scrape          re-fetch energy + royalty detail (3,762 GETs)
    python hydro.py latest          re-fetch only the current fiscal year (198 GETs)
    python hydro.py meta            re-fetch the plant register       (198 GETs)
    python hydro.py geo             re-fetch coordinates, districts, outline, terrain
    python hydro.py companies       rebuild the listed-company -> plant map
    python hydro.py mapsvg          export the fleet map as a standalone SVG
    python hydro.py map3d           MapLibre terrain map (local file, not an Artifact)
    python hydro.py all             scrape + meta + geo + build
    python hydro.py --stats         print the payload summary, write nothing

Everything lives in this one file. Page markup/CSS/JS is in template.html; this
substitutes the JSON payload into its __PAYLOAD__ placeholder.

All HTTP is GET only. The plant register's edit form is read but never submitted -
its POST action and CSRF token are deliberately untouched. Nominatim is called at
most once per second per its usage policy.
"""
import io, json, math, os, random, re, subprocess, sys, threading, time, unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed

import numpy as np
import pandas as pd
import requests
from bs4 import BeautifulSoup

HERE = os.path.dirname(os.path.abspath(__file__))
P = lambda *a: os.path.join(HERE, *a)
UA = {"User-Agent": "nepal-hydro-research/1.0 (public data reconciliation)"}

ENERGY_URL = "https://www.rmsdoed.gov.np/Core/Home/EnergyDetailIndex"
PLANT_URL  = "https://www.rmsdoed.gov.np/Core/Home/PlantCreate/"
XLSX       = os.environ.get("HYDRO_XLSX", r"D:\analysis\hydro\test2.xlsx")

NEP_MONTHS = ["Baisakh","Jestha","Asar","Shrawan","Bhadra","Aswin",
              "Kartik","Mangsir","Poush","Magh","Falgun","Chaitra"]
GREG = ["Apr–May","May–Jun","Jun–Jul","Jul–Aug","Aug–Sep","Sep–Oct",
        "Oct–Nov","Nov–Dec","Dec–Jan","Jan–Feb","Feb–Mar","Mar–Apr"]
FY = lambda x: 2000 + int(str(x).split('/')[0])      # "071/72" -> BS 2071
# Declared once, and it went stale the moment BS 2083 opened in April 2026: three
# plants commissioned that year showed an age of -1 on the map and the whole year
# fell out of the licensing chart, which filters on it. Derived from the data now;
# the constant is only the floor for a build with no monthly rows at all.
LATEST_BS_MIN = 2082


def latest_bs(d=None):
    """Newest Bikram Sambat year the data actually reaches."""
    if d is None or "BsYear" not in getattr(d, "columns", ()):
        return LATEST_BS_MIN
    v = pd.to_numeric(d.BsYear, errors="coerce").max()
    return max(LATEST_BS_MIN, int(v)) if np.isfinite(v) else LATEST_BS_MIN
DRY_MONTHS = {9, 10, 11, 12}                          # Poush-Chaitra carry the dry tariff


# ════════════════════════════════════════════════════════════ helpers ══════
def r(v, nd=2):
    """Round for JSON, mapping non-finite to None."""
    if v is None: return None
    try: v = float(v)
    except (TypeError, ValueError): return None
    return None if not np.isfinite(v) else round(v, nd)


def key(s):
    """Plant names differ between the two RMS pages by stray tabs, doubled
    spaces and case; join on a normalised key rather than the raw string."""
    return re.sub(r'\s+', ' ', str(s)).strip().lower()


def bs_year(v):
    """Bikram Sambat dates come back as YYYY/MM/DD on a few rows and DD/MM/YYYY
    on most, so take whichever component is a plausible 4-digit BS year rather
    than assuming a position. Guessing wrong silently voids every plant age."""
    if not isinstance(v, str): return np.nan
    for part in re.findall(r'\d+', v):
        if len(part) == 4 and 1990 <= int(part) <= 2200:
            return float(part)
    return np.nan


def num(s):
    """'1,336,087.09' -> float, elementwise, tolerant of blanks and dashes."""
    return pd.to_numeric(
        s.astype(str).str.replace(r"[,\s]", "", regex=True)
         .replace({"": None, "nan": None, "None": None, "-": None}), errors="coerce")


_local = threading.local()
def _sess():
    s = getattr(_local, "s", None)
    if s is None:
        s = requests.Session(); s.headers.update(UA)
        ad = requests.adapters.HTTPAdapter(pool_connections=16, pool_maxsize=16)
        s.mount("https://", ad); s.mount("http://", ad)
        _local.s = s
    return s


def _get(url, params=None, tries=3, timeout=45):
    last = None
    for k in range(tries):
        try:
            resp = _sess().get(url, params=params, timeout=timeout)
            resp.raise_for_status()
            return resp.text
        except Exception as e:
            last = e; time.sleep(1.5*(k+1) + random.random())
    raise last


def _retrieved(fname="rms_monthly.csv"):
    """When the register was last read, as a date string.

    From the commit that last touched the scrape, not the file's mtime. git does
    not preserve mtimes, so the mtime version reported the checkout date: every
    clone and every CI run claimed the data had been retrieved that morning. It
    also made the build non-deterministic, which is what stopped the commit step
    being able to judge the rebuilt page instead of enumerating input paths.
    """
    try:
        # A scrape that has just rewritten the file and not committed it yet was
        # retrieved now, not whenever it was last committed. Without this the date
        # lags a run behind, and the run after that rebuilds a page differing only
        # in the date and commits it -- an unchanged register reported as news,
        # every other day, forever.
        dirty = subprocess.run(["git", "-C", HERE, "diff", "--quiet", "--", fname],
                               capture_output=True, timeout=15).returncode
        if dirty:
            return time.strftime("%d %b %Y", time.gmtime())
        out = subprocess.run(["git", "-C", HERE, "log", "-1", "--format=%cs", "--", fname],
                             capture_output=True, text=True, timeout=15)
        stamp = out.stdout.strip()
        if stamp:
            return time.strftime("%d %b %Y", time.strptime(stamp, "%Y-%m-%d"))
    except Exception:
        pass
    return time.strftime("%d %b %Y", time.gmtime(os.path.getmtime(P(fname))))


def _pool(fn, jobs, workers, label):
    """Run fn over jobs, reporting progress; returns (results, failures)."""
    out, failed, done, t0 = [], [], 0, time.time()
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = {ex.submit(fn, *j if isinstance(j, tuple) else (j,)): j for j in jobs}
        for f in as_completed(futs):
            try: out.append(f.result())
            except Exception as e: failed.append((futs[f], repr(e)))
            done += 1
            if done % 250 == 0 or done == len(jobs):
                el = time.time() - t0
                print(f"  {label} {done}/{len(jobs)}  {el:.0f}s  fails={len(failed)}", flush=True)
    return out, failed


# ═══════════════════════════════════════════════════ scrape: energy detail ══
def dropdowns():
    soup = BeautifulSoup(_get(ENERGY_URL), "lxml")
    out = {}
    for sel in soup.find_all("select"):
        k = (sel.get("name") or sel.get("id") or "sel").lower()
        opts = {o.get("value"): o.get_text(strip=True)
                for o in sel.find_all("option") if o.get("value") not in (None, "0", "")}
        if opts: out[k] = opts
    return out


SUMMARY_LABELS = ("Previous Due", "Energy Royalty", "Capacity Royalty", "Received", "Balance")


def _flatten(cols):
    """read_html yields a 2-level MultiIndex; collapse it and strip the bare
    numbers the header row embeds (it splices each plant's Previous Due into a
    column name, which would otherwise fragment the concatenation)."""
    out, counts = [], {}
    for c in cols:
        seen = []
        for p in (c if isinstance(c, tuple) else (c,)):
            p = re.sub(r"\s+", " ", str(p)).strip()
            if p and not p.startswith("Unnamed") and p not in seen: seen.append(p)
        name = " ".join(t for t in " ".join(seen).split()
                        if not re.fullmatch(r"-?[\d,]+(?:\.\d+)?", t))
        name = re.sub(r"\s+", " ", name).strip(" .") or "col"
        counts[name] = counts.get(name, 0) + 1
        out.append(name if counts[name] == 1 else f"{name}_{counts[name]}")
    return out


def _summary_value(txt, label):
    """First occurrence actually followed by a number - a plain find() also hits
    nav text like 'Received Summary/Detail' and 'Balance Sheet'."""
    m = re.search(re.escape(label) + r"\s*:?\s*(?:Rs\.?)?\s*(-?[\d,]+(?:\.\d+)?)", txt)
    return m.group(1) if m else None


def scrape_one(fiscal_id, plant_id):
    html = _get(ENERGY_URL, {"FiscalId": fiscal_id, "PlantId": plant_id})
    soup = BeautifulSoup(html, "lxml")
    txt = soup.get_text(" ", strip=True)
    summ = {"PlantId": plant_id, "FiscalId": fiscal_id}
    for label in SUMMARY_LABELS:
        summ[label] = _summary_value(txt, label)
    try: tables = pd.read_html(io.StringIO(html))
    except ValueError: tables = []
    if not tables: return pd.DataFrame(), summ
    df = max(tables, key=len).copy()
    df.columns = _flatten(df.columns)
    df["PlantId"], df["FiscalId"] = plant_id, fiscal_id
    time.sleep(random.uniform(0.05, 0.30))
    return df, summ


def _merge_on(fresh, fname, keys, got, label):
    """Replace exactly the rows whose key is in `got`; leave every other row alone.

    Keyed on what actually came back, never on what was asked for. Writing only
    the response is the obvious thing and it is wrong: a request that fails
    contributes no rows, so its subject gets deleted outright. That applies to
    every scrape, not just the incremental one -- the full pass is the run that
    is supposed to repair the data, and it was the least protected of the three.

    Consequence worth naming: a row can only be removed by a successful fetch
    that returns nothing for it. A plant delisted from the site keeps its history
    here rather than evaporating. That is the intended direction of the failure.

    Existing rows are read as text and never coerced, so anything this pass did
    not touch is written back exactly as found. Round-tripping it through pandas'
    type inference would reformat numbers and rewrite the file for no reason.
    """
    path = P(fname)
    if fresh is None or not os.path.exists(path):
        return fresh
    old = pd.read_csv(path, dtype=str, keep_default_na=False, low_memory=False)
    if not set(keys) <= set(old.columns):
        return fresh
    key = list(zip(*[old[k].astype(str) for k in keys]))
    kept = old[[k not in got for k in key]]
    print(f"  {fname}: {len(kept):,} rows kept + {len(fresh):,} refreshed "
          f"({len(got):,} {label} re-read)")
    if not len(kept):
        return fresh
    return pd.concat([kept, fresh], ignore_index=True) if len(fresh) else kept


def cmd_scrape(workers=6, latest_only=False, years=2):
    d = dropdowns()
    plants  = d[[k for k in d if "plant" in k][0]]
    fiscals = d[[k for k in d if "fiscal" in k or "year" in k][0]]

    # A filed year never changes again, so re-fetching nineteen of them to learn
    # about one is the expensive way to find nothing. latest_only takes the newest
    # years and merges them over what is on disk: 396 requests instead of 3,762.
    #
    # Two years, not one, and the reason is the year boundary. The moment Shrawan
    # opens a new fiscal year the dropdown's newest entry moves, and the year that
    # just closed is still being filed -- its Jestha and Asar rows arrive weeks
    # late. Following only the newest entry would walk away from them.
    #
    # A new plant is picked up regardless: the plant list is re-read from the site
    # on every run, and a plant that has just started generating can only have
    # filings in these same newest years.
    keep_fids = None
    if latest_only:
        order = sorted(fiscals, key=lambda k: str(fiscals[k]))
        keep_fids = order[-years:]
        fiscals = {k: fiscals[k] for k in keep_fids}

    jobs = [(f, p) for p in plants for f in fiscals]
    print(f"{len(plants)} plants x {len(fiscals)} fiscal years = {len(jobs)} requests"
          + (f"  [latest {years}: {', '.join(fiscals.values())}]" if latest_only else ""))
    res, failed = _pool(scrape_one, jobs, workers, "energy")
    rows, summaries, got = [], [], set()
    for df, summ in res:
        pid, fid = summ["PlantId"], summ["FiscalId"]
        got.add((str(pid), str(fid)))        # what actually came back, for the merge
        if not df.empty:
            df["Plant"], df["FiscalYear"] = plants[pid], fiscals[fid]
            df["_seq"] = range(len(df))      # the site's own row order within a table
            rows.append(df)
        summ["Plant"], summ["FiscalYear"] = plants[pid], fiscals[fid]
        summaries.append(summ)

    # The pool yields in completion order, which varies run to run with six
    # threads on a variable network, so the raw files rewrote themselves almost
    # entirely on every scrape even when not one figure had changed. Sort to a
    # fixed key: identical data has to produce identical bytes, or a daily job
    # commits megabytes of reshuffling and reports it as news.
    monthly = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    summary = pd.DataFrame(summaries)

    # Always merge, full pass included. A failed request must not delete its
    # subject, and "we asked for everything" is not the same as "everything came
    # back" -- CI has returned nothing for half a run.
    K = ["PlantId", "FiscalId"]
    monthly = _merge_on(monthly, "rms_monthly.csv", K, got, "plant-years")
    summary = _merge_on(summary, "rms_summary.csv", K, got, "plant-years")

    if not monthly.empty:
        # rows recovered from disk carry no _seq; they sort last within their
        # key, which is harmless because a key is never half-old and half-fresh
        if "_seq" not in monthly.columns:
            monthly["_seq"] = np.nan
        monthly = (monthly.sort_values(["PlantId", "FiscalId", "_seq"], kind="stable")
                          .drop(columns="_seq").reset_index(drop=True))
    monthly.to_csv(P("rms_monthly.csv"), index=False)
    if not summary.empty:
        summary = (summary.sort_values(["PlantId", "FiscalId"], kind="stable")
                          .reset_index(drop=True))
    summary.to_csv(P("rms_summary.csv"), index=False)
    print(f"monthly {monthly.shape} | failures {len(failed)}")
    clean()


def clean():
    """Normalise the raw scrape into the analysis tables the build reads."""
    m = pd.read_csv(P("rms_monthly.csv"), low_memory=False)
    # the scraped table has the page's summary footer appended to it; real rows
    # are exactly those whose "Year Month" is a BS yyyy/m stamp
    ym = m["Year Month"].astype(str).str.extract(r"^(\d{4})\s*/\s*(\d{1,2})$")
    d = m[ym[0].notna()].copy()
    d["BsYear"]  = ym.loc[d.index, 0].astype(int)
    d["BsMonth"] = ym.loc[d.index, 1].astype(int)
    d = d[d.BsMonth.between(1, 12)]              # month 13 is an all-NaN padding row

    col = lambda n: num(d[n]) if n in d.columns else pd.Series(np.nan, index=d.index)
    # Total Energy Generation is blank on 188 rows, but the site also publishes its
    # components and they reconcile exactly on 96.5% of rows where both are present
    # (Total = Metered + Additional + Internal consumption + Transmission loss), so
    # rebuild the total rather than lose the month.
    _parts = pd.concat([col("Metered Energy"), col("Additional Energy"),
                        col("Internal Consumption"), col("Transmission Loss")], axis=1)
    d["Generation_kWh"] = col("Total Energy Generation").fillna(
                              _parts.sum(axis=1, min_count=1))
    d["Metered_kWh"]    = col("Metered Energy")
    # The site serves three schemas. Two are plain NPR; the third bills in USD and
    # must be converted and added to any NPR portion of the same invoice - skipping
    # it silently drops the entire revenue line for foreign-currency PPAs such as
    # Upper Bhotekoshi. Checked against the statutory royalty rate: this lands
    # 146/158 USD rows within 0.05pp of an exact 2% or 10%.
    fx  = col("Exchange Rate")
    _usd_raw = col("Invoice to NEA(USD)").fillna(col("Amount paid by NEA (USD)"))
    _usd_raw = _usd_raw.where(_usd_raw > 0)      # a filed zero means "not billed in USD
                                                 # this month", not "revenue was zero"
    usd = _usd_raw * fx
    usd = usd + col("Invoice to NEA(NRS)").where(usd.notna()).fillna(0)
    # USD first: on a mixed invoice "Invoice to NEA(NRS)" holds only the NPR
    # portion, so letting it win would report a fraction of the month's revenue
    # (Upper Bhotekoshi showed NPR 11.8m against a USD 2.94m invoice).
    d["Revenue_NPR"]    = (usd
                           .fillna(col("Billed Amount to NEA"))
                           .fillna(col("Invoice to NEA(NRS)"))
                           .fillna(col("Amount paid by NEA Rs"))
                           .fillna(col("Amount paid by NEA (NRs)"))
                           .fillna(col("Revenue From R.E (NRs) Local Distribution")))
    d["Royalty_NPR"]    = col("Royalty Amount")
    # foreign-currency PPAs are contracted in USD; keep the filed USD figures so the
    # page can report them natively instead of only as a reconstructed NPR number
    d["Revenue_USD"]    = _usd_raw
    d["FX"]             = fx
    # "Seasional Rate (USD)" is a mixed column: values below 1 really are USD/kWh
    # (0.00-0.12), while values at or above 1 are Nepali NPR tariffs (4.80, 5.09,
    # 5.23, 8.40, 9.16) mis-filed into the USD field. Split on magnitude - the two
    # populations do not overlap - and route each to the right column.
    _rate_usd_col = col("Seasional Rate (USD)")
    d["Rate_USD_kWh"]   = _rate_usd_col.where(_rate_usd_col.between(0, 1, inclusive="neither"))
    d["Rate_NPR_kWh"]   = (col("Seasional Rate").fillna(col("Rate to NEA"))
                                                .fillna(_rate_usd_col.where(_rate_usd_col >= 1)))
    d["Capacity_kW"]    = num(d["Plant"].str.extract(r"([\d.]+)\s*kW", flags=re.I)[0])
    d["PlantName"] = d["Plant"].str.replace(r"\s*,\s*[\d.]+\s*kW\s*$", "", regex=True).str.strip()
    d["MonthName"] = d.BsMonth.map(lambda i: NEP_MONTHS[i-1])
    d["Period"]    = d.BsYear.astype(str) + "/" + d.BsMonth.astype(str).str.zfill(2)
    d["CapacityFactor"] = d.Generation_kWh / (d.Capacity_kW * 730.0)

    # some NEA plants file a whole year as ONE row; those carry an impossible
    # monthly capacity factor, so flag them out of seasonality but keep the total
    impCF = d.CapacityFactor.copy()
    per_year = d.groupby(["PlantName", "FiscalYear"])["Period"].transform("size")
    d["IsAnnualFiling"] = ((per_year <= 3) & (impCF > 1.3)) | (impCF > 2.0)
    d.loc[d.IsAnnualFiling | ~np.isfinite(d.CapacityFactor) | (d.CapacityFactor > 1.5),
          "CapacityFactor"] = np.nan

    keep = ["PlantId","PlantName","Capacity_kW","FiscalId","FiscalYear","BsYear","BsMonth",
            "MonthName","Period","Generation_kWh","Metered_kWh","Revenue_NPR","Royalty_NPR",
            "Rate_NPR_kWh","CapacityFactor","IsAnnualFiling",
            "Revenue_USD","Rate_USD_kWh","FX"]
    d[keep].sort_values(["PlantName","BsYear","BsMonth"]).to_csv(
        P("rms_monthly_clean.csv"), index=False)

    s = pd.read_csv(P("rms_summary.csv"), low_memory=False)
    s = s[s["Energy Royalty"].notna()].copy()
    for c in SUMMARY_LABELS:
        s[c.replace(" ", "_")] = num(s[c])
    s["Capacity_kW"] = num(s["Plant"].str.extract(r"([\d.]+)\s*kW", flags=re.I)[0])
    s["PlantName"] = s["Plant"].str.replace(r"\s*,\s*[\d.]+\s*kW\s*$", "", regex=True).str.strip()
    s["RoyaltyDue"] = s.Energy_Royalty.fillna(0) + s.Capacity_Royalty.fillna(0)
    s[["PlantId","PlantName","Capacity_kW","FiscalId","FiscalYear","Previous_Due",
       "Energy_Royalty","Capacity_Royalty","RoyaltyDue","Received","Balance"]].to_csv(
        P("rms_summary_clean.csv"), index=False)
    print(f"cleaned: {len(d)} monthly rows, {len(s)} plant-year summaries")


# ════════════════════════════════════════════════════ scrape: plant register ══
META_TEXT = ["Id","PlantName","LicenseNumber","PlantCapacity","LicensedMiti",
             "LicenseExpiryMiti","MitiofOperation","WardNo","VDCMNC","PhoneNo","Email",
             "URL","Rivers","Latitude","Longitude","DateofOperation","LicensedDate",
             "LicenseExpiryDate","CreatedBy","CreatedDate","IsActive"]
META_SELECT = ["CompanyId","PlantType","DistrictId","ProvinceId"]


def scrape_plant(pid):
    s = BeautifulSoup(_get(PLANT_URL + str(pid)), "lxml")
    out = {"UrlId": pid}
    for n in META_TEXT:
        e = s.find(attrs={"name": n})
        out[n] = e.get("value") if e is not None else None
    for n in META_SELECT:
        e = s.find(attrs={"name": n})
        o = e.find("option", selected=True) if (e is not None and e.name == "select") else None
        out[n] = o.get_text(strip=True) if o else None
    time.sleep(random.uniform(0.05, 0.25))
    return out


def cmd_meta(workers=6):
    ids = sorted(int(k) for k in dropdowns()['plantid'])
    print(f"plant register: {len(ids)} pages")
    rows, failed = _pool(scrape_plant, ids, workers, "meta")
    fresh = pd.DataFrame(rows)
    # Same rule as the energy scrape: a page that failed must keep the plant it
    # describes. Writing only the response would drop it, taking its commissioning
    # date, company, district and map position with it, and ten could go before
    # the row-count guard noticed.
    got = {(str(r["UrlId"]),) for r in rows}
    merged = _merge_on(fresh, "rms_plants_meta.csv", ["UrlId"], got, "plants")
    merged = merged.copy()
    merged["UrlId"] = pd.to_numeric(merged.UrlId, errors="coerce")
    merged.sort_values("UrlId").to_csv(P("rms_plants_meta.csv"), index=False)
    print(f"wrote rms_plants_meta.csv ({len(merged)} plants, {len(failed)} failures)")


# ═════════════════════════════════════════════════════════════ geo lookups ══
OVERPASS = ["https://overpass-api.de/api/interpreter",
            "https://overpass.kumi.systems/api/interpreter"]
OVERPASS_Q = """
[out:json][timeout:180];
area["ISO3166-1"="NP"][admin_level=2]->.np;
( node["power"="plant"]["plant:source"="hydro"](area.np);
  way["power"="plant"]["plant:source"="hydro"](area.np);
  relation["power"="plant"]["plant:source"="hydro"](area.np);
  node["power"="generator"]["generator:source"="hydro"](area.np);
  way["power"="generator"]["generator:source"="hydro"](area.np); );
out center tags;"""
Q_RIVERS = """
[out:json][timeout:280];
area["ISO3166-1"="NP"][admin_level=2]->.np;
( way["waterway"="river"]["name"](area.np); );
out geom;"""
# Most of the fleet sits on a khola, not on a main stem, and OSM tags those
# waterway=stream - so they are absent from Q_RIVERS at any length threshold.
# Nepal has ~1,460 named streams; only the ones a plant is actually named for
# are kept (see _fleet_keys), which is what makes embedding them affordable.
Q_STREAMS = """
[out:json][timeout:280];
area["ISO3166-1"="NP"][admin_level=2]->.np;
( way["waterway"="stream"]["name"](area.np); );
out geom;"""
Q_PEAKS = """
[out:json][timeout:180];
area["ISO3166-1"="NP"][admin_level=2]->.np;
( node["natural"="peak"]["ele"](area.np); );
out tags center;"""

Q_GRID = """
[out:json][timeout:280];
area["ISO3166-1"="NP"][admin_level=2]->.np;
( way["power"="line"](area.np); );
out geom tags;"""
Q_SUBS = """
[out:json][timeout:240];
area["ISO3166-1"="NP"][admin_level=2]->.np;
( node["power"="substation"](area.np); way["power"="substation"](area.np); );
out tags center;"""

WIKIDATA_Q = """
SELECT ?item ?itemLabel ?coord WHERE {
  ?item wdt:P17 wd:Q837 .
  { ?item wdt:P31/wdt:P279* wd:Q15911738 } UNION { ?item wdt:P31/wdt:P279* wd:Q11165949 }
  ?item wdt:P625 ?coord .
  SERVICE wikibase:label { bd:serviceParam wikibase:language "en,ne" } }"""


def _nominatim(q, poly=False):
    p = {"q": q, "format": "json", "limit": 1, "countrycodes": "np"}
    if poly: p["polygon_geojson"] = 1
    resp = requests.get("https://nominatim.openstreetmap.org/search",
                        params=p, headers=UA, timeout=90)
    time.sleep(1.1)                                   # policy: <= 1 req/second
    resp.raise_for_status()
    j = resp.json()
    return j[0] if j else None


def _rdp(pts, eps):
    if len(pts) < 3: return pts
    def perp(p, a, b):
        (x, y), (x1, y1), (x2, y2) = p, a, b
        dx, dy = x2-x1, y2-y1
        if dx == 0 and dy == 0: return math.hypot(x-x1, y-y1)
        t = max(0, min(1, ((x-x1)*dx + (y-y1)*dy)/(dx*dx + dy*dy)))
        return math.hypot(x-(x1+t*dx), y-(y1+t*dy))
    dmax, idx = 0.0, 0
    for i in range(1, len(pts)-1):
        dd = perp(pts[i], pts[0], pts[-1])
        if dd > dmax: dmax, idx = dd, i
    if dmax > eps:
        return _rdp(pts[:idx+1], eps)[:-1] + _rdp(pts[idx:], eps)
    return [pts[0], pts[-1]]


def _overpass(q):
    for url in OVERPASS:
        try:
            resp = requests.post(url, data={"data": q}, headers=UA, timeout=300)
            if resp.status_code == 200: return resp.json()
            print(f"  overpass {url} -> {resp.status_code}")
        except Exception as e:
            print(f"  overpass {url} -> {type(e).__name__}")
    return None


def _seglen(pts):
    """Rough length in km, lon compressed by Nepal's mid-latitude."""
    tot = 0.0
    for i in range(1, len(pts)):
        dx = (pts[i][0]-pts[i-1][0]) * math.cos(math.radians(28.4)) * 111.32
        dy = (pts[i][1]-pts[i-1][1]) * 110.57
        tot += math.hypot(dx, dy)
    return tot


def _ways_by_name(j):
    """Overpass returns a watercourse split into many ways sharing one name."""
    from collections import defaultdict
    by_name = defaultdict(list)
    for w in (j or {}).get("elements", []):
        if not w.get("geometry"): continue
        nm = (w.get("tags", {}).get("name:en") or w.get("tags", {}).get("name") or "").strip()
        if not nm: continue
        pts = [(round(g["lon"], 4), round(g["lat"], 4)) for g in w["geometry"]]
        if len(pts) >= 2: by_name[nm].append(pts)
    return by_name


def _fleet_keys():
    """Every normalised name the fleet could be looked up by - the river the
    register names for a plant, and the plant's own name, since a station is
    almost always named for the water it sits on. Single tokens are included
    because _snap_to_river falls back to them."""
    try:
        m = pd.read_csv(P("rms_plants_meta.csv"), dtype=str)
    except FileNotFoundError:
        return set()
    keys = set()
    for col in ("Rivers", "PlantName"):
        if col not in m.columns: continue
        for v in m[col].dropna():
            core = _parts(v)[0]
            if not core: continue
            keys.add(frozenset(core))
            keys |= {frozenset([t]) for t in core}
    return keys


def fetch_terrain(min_km=25, min_ele=6800):
    """Nepal's river network and high peaks - the physical context the fleet sits in.
    Rivers arrive split into many ways sharing a name, so group by name, drop the
    creeks, and simplify hard enough to embed.

    Three Overpass calls, and the API rate-limits. A part whose call fails keeps
    whatever the last good run wrote rather than overwriting it with nothing --
    losing the peaks because the rivers used up the quota is not an improvement."""
    sys.setrecursionlimit(30000)
    try:
        prev = json.load(open(P("np_terrain.json"), encoding="utf-8"))
    except (FileNotFoundError, ValueError):
        prev = {}
    kept = []

    jr = _overpass(Q_RIVERS)
    if jr is None:
        rivers, _ = prev.get("rivers", []), kept.append("rivers")
    else:
        rivers = []
        for nm, segs in _ways_by_name(jr).items():
            km = sum(_seglen(x) for x in segs)
            if km < min_km: continue
            simp = [p for p in (_rdp(x, 0.008) for x in segs) if len(p) >= 2]
            if simp: rivers.append({"n": nm, "km": round(km), "segs": simp})
        rivers.sort(key=lambda x: -x["km"])

    # ── the tributaries the fleet is actually built on ──────────────────────
    # A main stem carries a handful of stations; the rest are on kholas that
    # Q_RIVERS never returns. Keeping only fleet-named streams turns ~1,460
    # candidates into a few dozen, small enough to embed and snap against.
    want = _fleet_keys()
    have = {frozenset(_parts(r["n"])[0]) for r in rivers}
    js = _overpass(Q_STREAMS) if want else {}
    if js is None:
        streams, _ = prev.get("streams", []), kept.append("streams")
    else:
        streams = []
        for nm, segs in _ways_by_name(js).items():
            k = frozenset(_parts(nm)[0])
            if not k or k in have or k not in want: continue
            km = sum(_seglen(x) for x in segs)
            # a stream simplified as hard as a main stem loses its shape
            simp = [p for p in (_rdp(x, 0.0015) for x in segs) if len(p) >= 2]
            if simp: streams.append({"n": nm, "km": round(km), "segs": simp, "sm": 1})
        streams.sort(key=lambda x: -x["km"])

    j2 = _overpass(Q_PEAKS)
    if j2 is None:
        peaks, _ = prev.get("peaks", []), kept.append("peaks")
    else:
        raw = []
        for e in j2.get("elements", []):
            t = e.get("tags", {})
            nm = t.get("name:en") or t.get("name")
            mm = re.match(r"^\s*(\d{3,5})", str(t.get("ele", "")))
            if not nm or not mm or e.get("lat") is None: continue
            h = int(mm.group(1))
            if h >= min_ele:
                raw.append({"n": nm, "e": h, "la": round(e["lat"], 4), "lo": round(e["lon"], 4)})
        # OSM maps subsidiary summits separately (Lhotse Shar, Everest South Peak...);
        # keep the highest point of each massif so the map reads as mountains, not noise
        raw.sort(key=lambda p: -p["e"])
        peaks = []
        for p in raw:
            if any(abs(p["la"]-q["la"]) < 0.06 and abs(p["lo"]-q["lo"]) < 0.06 for q in peaks):
                continue
            peaks.append(p)

    json.dump({"rivers": rivers, "streams": streams, "peaks": peaks},
              open(P("np_terrain.json"), "w", encoding="utf-8"), ensure_ascii=False)
    if kept: print(f"  overpass failed for {', '.join(kept)} - kept the previous data")
    print(f"terrain: {len(rivers)} rivers ({sum(len(x) for r in rivers for x in r['segs'])} pts),"
          f" {len(streams)} fleet tributaries"
          f" ({sum(len(x) for r in streams for x in r['segs'])} pts),"
          f" {len(peaks)} peaks")


def _volts(tag):
    """OSM voltage is in volts and may list several ('132000;66000'); take the
    highest, in kV. Nepal runs 400 / 220 / 132 / 66 / 33 kV."""
    best = None
    for part in re.findall(r"\d+", str(tag or "")):
        v = int(part)
        if 1000 <= v <= 800000:
            kv = round(v / 1000)
            best = kv if best is None else max(best, kv)
    return best


def fetch_grid(min_kv=33):
    """Nepal's transmission network - the grid the fleet has to reach.
    The RPGCL network map (rpgcl.com) is the authoritative picture but is a
    rendered PDF with no coordinates, so geometry comes from OpenStreetMap."""
    sys.setrecursionlimit(30000)
    j = _overpass(Q_GRID)
    lines = []
    for w in (j or {}).get("elements", []):
        if not w.get("geometry"): continue
        t = w.get("tags", {})
        kv = _volts(t.get("voltage"))
        if kv is None or kv < min_kv: continue
        pts = [(round(g["lon"], 4), round(g["lat"], 4)) for g in w["geometry"]]
        pts = _rdp(pts, 0.004) if len(pts) > 2 else pts
        if len(pts) < 2: continue
        lines.append({"kv": kv, "n": (t.get("name") or t.get("ref") or "").strip() or None,
                      "pts": pts})
    lines.sort(key=lambda x: -x["kv"])

    j2 = _overpass(Q_SUBS)
    subs, seen = [], set()
    for e in (j2 or {}).get("elements", []):
        t = e.get("tags", {})
        nm = (t.get("name:en") or t.get("name") or "").strip()
        c = e.get("center") or {"lat": e.get("lat"), "lon": e.get("lon")}
        if not nm or c.get("lat") is None: continue
        k = key(nm)
        if k in seen: continue
        seen.add(k)
        subs.append({"n": nm, "kv": _volts(t.get("voltage")),
                     "la": round(c["lat"], 4), "lo": round(c["lon"], 4)})
    subs.sort(key=lambda x: -(x["kv"] or 0))
    json.dump({"lines": lines, "subs": subs},
              open(P("np_grid.json"), "w", encoding="utf-8"), ensure_ascii=False)
    from collections import Counter
    print(f"grid: {len(lines)} lines ({sum(len(l['pts']) for l in lines)} pts), "
          f"{len(subs)} substations")
    print("  by kV:", dict(sorted(Counter(l["kv"] for l in lines).items(), reverse=True)))


def cmd_geo():
    # 1. plant coordinates from OpenStreetMap and Wikidata
    cands = []
    for url in OVERPASS:
        try:
            resp = requests.post(url, data={"data": OVERPASS_Q}, headers=UA, timeout=200)
            if resp.status_code != 200: continue
            for e in resp.json().get("elements", []):
                t = e.get("tags", {}); nm = t.get("name:en") or t.get("name")
                c = e.get("center") or {"lat": e.get("lat"), "lon": e.get("lon")}
                if nm and c.get("lat") is not None:
                    cands.append({"name": nm, "lat": float(c["lat"]),
                                  "lon": float(c["lon"]), "src": "OSM"})
            break
        except Exception as e:
            print(f"  overpass {url}: {type(e).__name__}")
    try:
        resp = requests.get("https://query.wikidata.org/sparql",
                            params={"query": WIKIDATA_Q, "format": "json"},
                            headers=UA, timeout=120)
        for b in resp.json()["results"]["bindings"]:
            mm = re.match(r"Point\(([-\d.]+) ([-\d.]+)\)", b["coord"]["value"])
            if mm: cands.append({"name": b["itemLabel"]["value"], "lat": float(mm.group(2)),
                                 "lon": float(mm.group(1)), "src": "Wikidata"})
    except Exception as e:
        print("  wikidata:", type(e).__name__)
    seen, uniq = set(), []
    for c in cands:
        k = key(c["name"])
        if k not in seen: seen.add(k); uniq.append(c)
    json.dump(uniq, open(P("geo_candidates.json"), "w", encoding="utf-8"), ensure_ascii=False)
    print(f"coordinate candidates: {len(uniq)}")

    # 2. district centroids
    m = pd.read_csv(P("rms_plants_meta.csv"), dtype=str).replace(r"^\s*$", np.nan, regex=True)
    ALIAS = {"Rashuwa":"Rasuwa","Bajhanga":"Bajhang","Okhaldunga":"Okhaldhunga",
             "Tehrathum":"Terhathum","Makawanpur":"Makwanpur","Rukum":"Rukum East"}
    dists = {}
    names = sorted(m.DistrictId.dropna().unique())
    for i, dn in enumerate(names, 1):
        q = ALIAS.get(dn, dn)
        hit = _nominatim(f"{q} District, Nepal") or _nominatim(f"{q}, Nepal")
        if hit:
            dists[dn] = {"lat": float(hit["lat"]), "lon": float(hit["lon"])}
        print(f"  district {i}/{len(names)} {dn}", flush=True)
    json.dump(dists, open(P("np_districts.json"), "w", encoding="utf-8"), ensure_ascii=False)

    # 3. national outline
    hit = _nominatim("Nepal", poly=True)
    if hit and "geojson" in hit:
        g = hit["geojson"]; c = g["coordinates"]
        rings = c if g["type"] == "Polygon" else [rr for poly in c for rr in poly]
        out = []
        for ring in rings:
            simp = _rdp([(round(x, 4), round(y, 4)) for x, y in ring], 0.010)
            if len(simp) >= 8: out.append(simp)
        out.sort(key=len, reverse=True); out = out[:3]
        xs = [p[0] for rr in out for p in rr]; ys = [p[1] for rr in out for p in rr]
        json.dump({"rings": out, "bbox": [min(xs), min(ys), max(xs), max(ys)]},
                  open(P("np_outline.json"), "w", encoding="utf-8"))
        print(f"outline: {sum(len(rr) for rr in out)} points")
    fetch_terrain()
    fetch_grid()
    resolve_coords()


# ── strict name matching ──────────────────────────────────────────────────────
NOISE = {"hydropower","hydro","power","plant","project","station","hpp","hp","hep","shpp",
         "the","electricity","company","ltd","pvt","limited","generation","house","of","co",
         "hydroelectric","hydel","hydropwer","khola","nadi","gad","gadh","river","dam","unit","mw"}
QUALIFIERS = {"upper","lower","middle","madhya","mathillo","tallo","super","sano","thulo",
              "chhoto","small","mini","micro","big","main","cascade","beni",
              "i","ii","iii","iv","v","vi","1","2","3","4","5","6",
              "a","b","c","d","3a","1a","2a","kha","ka","ga"}
TRANSLIT = [("sh","s"),("chh","ch"),("ph","f"),("aa","a"),("ee","i"),("oo","u"),("w","v")]


def _parts(s):
    if not isinstance(s, str): return set(), set()
    s = unicodedata.normalize("NFKD", s).lower()
    s = re.sub(r"[^a-z0-9\s]", " ", s.replace("'", " ").replace("-", " ").replace("_", " "))
    toks = [t for t in s.split() if t]
    qual = {t for t in toks if t in QUALIFIERS}
    core = set()
    for t in toks:
        if t in QUALIFIERS or t in NOISE: continue
        for a, b in TRANSLIT: t = t.replace(a, b)
        core.add(re.sub(r"(.)\1+", r"\1", t))
    return core, qual


def _is_match(a, b):
    """Sibling stations differ by a single qualifier - Upper/Lower, an ordinal, a
    size class - so 'Kulekhani-II' and 'Kulekhani III' are different plants that a
    fuzzy scorer happily merges. Require core AND qualifiers to agree."""
    ca, qa = _parts(a); cb, qb = _parts(b)
    if not ca or not cb or qa != qb: return False
    if ca == cb: return True
    if len(ca) == 1 and len(cb) == 1:            # tolerate one transposition on a long core
        x, y = next(iter(ca)), next(iter(cb))
        if len(x) >= 8 and len(y) >= 8 and abs(len(x)-len(y)) <= 1 and sorted(x) == sorted(y):
            return True
    return False


def _river_index():
    """Map a watercourse's core-name token set -> its vertices, for snapping.
    Tributaries count: a plant named for a khola snaps far closer on that khola
    than on the main stem it eventually joins."""
    try:
        terr = json.load(open(P("np_terrain.json"), encoding="utf-8"))
    except FileNotFoundError:
        return {}
    idx = {}
    for rv in terr.get("rivers", []) + terr.get("streams", []):
        k = frozenset(_parts(rv["n"])[0])
        if not k: continue
        pts = [p for seg in rv["segs"] for p in seg]
        e = idx.get(k)
        if e is None:
            idx[k] = [rv["n"], pts, rv["km"], rv["km"]]   # name, pts, total km, best variant
        else:
            # same river under another spelling: one watercourse, so pool the
            # vertices and the length, and label it with the longest variant
            e[1] = e[1] + pts
            e[2] += rv["km"]
            if rv["km"] > e[3]: e[0], e[3] = rv["n"], rv["km"]
    return idx


def _named_rivers(field):
    """The register sometimes names two watercourses in one field -- 'Kabeli
    Khola, Amji Khola' for a plant that draws from both. Yield them in the order
    written, so the one the register puts first is the one tried first; reducing
    the whole field to a token set loses that and picks arbitrarily."""
    if not isinstance(field, str): return []
    return [p for p in re.split(r"[,/;&]| and ", field) if p.strip()]


def _snap_to_river(plant_name, river_field, lat, lon, idx, ceil_km=30,
                   taken=None, sep_km=0.8):
    """A district centroid is a town or a ridge; a hydropower plant is on a
    watercourse. Where the register names the river (or the plant is named after
    it) and that river is mapped, move the marker to the point on THAT river
    nearest the district centre. Still approximate - but on the right river, and
    constrained to the right district.

    How far the marker may travel is bounded by the watercourse's own mapped
    length, because khola names repeat all over Nepal: a 5 km Pikhuwa Khola found
    39 km from the district it is supposed to be in is a different Pikhuwa Khola,
    and a district centroid drawn as a hollow dot is a more honest answer than a
    filled one on the wrong stream."""
    hit = None
    for src in _named_rivers(river_field) + [plant_name]:
        if not isinstance(src, str): continue
        core = _parts(src)[0]
        if not core: continue
        k = frozenset(core)
        if k in idx: hit = idx[k]; break
        for tok in sorted(core):                 # sorted: never depend on set order
            kk = frozenset([tok])
            if kk in idx: hit = idx[kk]; break
        if hit: break
    if not hit: return None
    name, pts, km = hit[0], hit[1], hit[2]
    max_km = min(ceil_km, max(10, km))
    kx = math.cos(math.radians(28.4))
    cand = sorted((math.hypot((lo-lon)*kx*111.32, (la-lat)*110.57), lo, la) for lo, la in pts)
    cand = [c for c in cand if c[0] <= max_km]
    if not cand: return None
    # A cascade files every station under one district on one river, and each
    # would otherwise take the same nearest vertex and draw as a single dot that
    # only one of them could be clicked on. The point along the river was always
    # arbitrary, so spreading them down the reach costs no accuracy and is the
    # difference between six plants being visible and one being visible.
    used = None if taken is None else taken.setdefault(name, [])
    if used is not None:
        for d, lo, la in cand:
            if any(math.hypot((lo-ulo)*kx*111.32, (la-ula)*110.57) < sep_km
                   for ula, ulo in used): continue
            used.append((la, lo))
            return la, lo, name
    _, lo, la = cand[0]                      # reach full: fall back to the nearest
    if used is not None: used.append((la, lo))
    return la, lo, name


# The register's DistrictId is sometimes the company's registered office rather
# than where the plant stands. It is rare -- two rows of 198 -- but it moves a
# plant the width of the country on the map, and the river snap cannot rescue it
# because the named river is then hundreds of km from the centroid it starts at.
# Corrected by hand, with the evidence, in the same spirit as the ALIAS table.
DISTRICT_FIX = {
    # Tundi Power's Chimkhola-Rahughat-Mangale, on the Rahughat in Raghuganga RM,
    # Myagdi. The register files it under Lalitpur, which is the office.
    "Rahughat Mangale": "Myagdi",
    # 21.3 MW at Kuinemangle, Raghuganga RM-8, Myagdi -- the same municipality.
    "Thulo Khola HPP":  "Myagdi",
}


# Positions the author looked up on Google Maps, one plant at a time, and gave
# as plus codes. They are the only locations here that a person checked against
# satellite imagery rather than a name match, so they outrank both the candidate
# match and the river snap. Each was decoded against its district centre and
# then verified: inside Nepal, inside a plausible radius of that district, and -
# where OpenStreetMap names the watercourse the register gives - close to it.
# Two rows of the batch were dropped, both a copy of the row above's code:
# "Solu" carried Mathillo Thulo Khola A's point at Chimkhola in Myagdi, 87 km
# from any Solu Khola, and "Sipring Khola HP" carried Upper Chaku A's point at
# Phulpingkatti in Sindhupalchok, 40 km outside Dolakha.
# name -> (lat, lon, plus code, the locality Google printed with it)
PLUS_FIX = {
    "Ankhukhola Small HP": (27.999938, 84.931953, "XWXJ+XQG", "45100"),
    "Baramchi Khola HPP": (27.836137, 85.775859, "RQPG+F85", "Baramchi 45308"),
    "Bijaypur-1 Small HP": (28.178887, 84.033078, "52HM+H64", "Lekhnath 33700"),
    "Buku Khola HPP": (27.523687, 86.350812, "G9F2+F8", "Rawadolu"),
    "Buku-Kapati HPP": (27.518512, 86.398641, "G99X+CF2", "Goli 56000"),
    "Chaku Khola  Small HP": (27.875512, 85.929516, "VWGH+6R2", ""),
    "Chatara HP": (26.815963, 87.157203, "R584+9VP", "Barahakshetra 56703"),
    "Dordi 1": (28.230812, 84.450688, "6FJ2+87", "Chiti 33600"),
    "Dordi Khola": (28.170063, 84.442484, "5CCR+2XH", "Unnamed Road, Chiti 33600"),
    "Down Piluwa": (27.245188, 87.280438, "67WJ+35", "Ankhibhui 56900"),
    "Fewa (Pokhara)  HP": (28.182163, 83.972234, "5XJC+VV9", "Unnamed Road, Pokhara 33700"),
    "Ghalemdi Khola": (28.536887, 83.675078, "GMPG+Q24", "Narchyang 33200"),
    "Ghar Khola": (28.479337, 83.642922, "FJHV+P5M", "Pokharebagar 33200"),
    "Ghatte Khola": (27.786937, 86.286812, "Q7PP+QP", "Sikpasor Khani 45500"),
    "Hidi Khola HEP": (28.226562, 84.110062, "64G6+J2", "Saimarang 33700"),
    "Jiri Khola SHP": (27.585663, 86.226672, "H6PG+7M7", "Jiri 45500"),
    "Jogmai Cascade HPP": (26.897738, 87.994422, "VXXV+3QV", "Namsaling 57300"),
    "Jogmai Khola": (26.913937, 88.018797, "W279+HGF", "Namsaling 57300"),
    "Kapadigad HPP": (29.004437, 80.759187, "2Q35+QM", "BP Nagar 10800"),
    "Khorunga Khola HPP": (27.159987, 87.544641, "5G5V+XVR", "Oyakjung 57500"),   # shared point
    "Kulekhani-I  HP": (27.540412, 85.114172, "G4R7+5M7", "Bhimphedi 44100"),
    "Kulekhani-II  HP": (27.512688, 85.047828, "G27X+34G", "Bhaise 44100"),
    "Lankhuwa Khola HPP": (27.290312, 87.193938, "75RV+4H", "Heel 33300"),
    "Lower Chaku HP": (27.882813, 85.910672, "VWM6+47F", "Chaku"),
    "Lower Jogmai Khola HPP": (26.890687, 87.959437, "VXR5+7Q", "Namsaling 57300"),
    "Lower Khare": (27.758538, 86.197172, "Q55W+CV7", "Charikot - Lamabagar Rd, Sunkhani 45500"),
    "Lower Tadi Khola": (27.922837, 85.349734, "W8FX+4VQ", "Thaprek 44900"),
    "Madme Khola HPP": (28.419562, 84.116312, "C498+RG", "Namarjung 33700"),
    "Mailung Khola": (28.072513, 85.209172, "36F5+2M3", "Mailung Bazzar Bus Stop, Dandagaun 45000"),
    "Mathillo Thulo Khola A HEP": (28.581937, 83.537562, "HGJQ+Q2", "Chimkhola 33200"),
    "Mewa Khola HPP": (27.416162, 87.614484, "CJ87+FQ9", "57500"),
    "Middle Chaku HP": (27.872138, 85.933297, "VWCM+V83", ""),
    "Mistry Khola": (28.536962, 83.674422, "GMPF+QQM", "Narchyang 33200"),   # shared point
    "Nilgiri Khola ": (28.578712, 83.692359, "HMHR+FWQ", "Narchyang 33200"),
    "Nilgiri Khola-II cascade Project": (28.536962, 83.674422, "GMPF+QQM", "Narchyang 33200"),   # shared point
    "Phawa khola": (27.280888, 87.756828, "7QJ4+9P4", "Unnamed Road, Thechambu 57500"),
    "Pikhuwa Khola": (27.120338, 87.066016, "43C8+4CJ", "Bhojpur 57000"),   # shared point
    "Radhi Small": (28.396613, 84.429172, "9CWH+JMV", "33600"),
    "Rahughat Mangale": (28.475562, 83.518422, "FGG9+69F", "Sisneri, Pakhapani 33200"),
    "Rawa Khola": (27.348187, 86.840063, "8RXR+72", "Sungdel 56200"),
    "Rele Khola HPP": (28.530188, 83.679813, "GMJH+3W", "Narchyang 33200"),
    "Rudi A": (28.231262, 84.210547, "66J6+G63", "Madi-9, Baluwabesi"),
    "Sagu Khola HEP": (27.788113, 86.115547, "Q4Q8+66V", "45500"),
    "Sanjen": (28.217937, 85.284312, "679M+5P", "Chilime"),
    "Sapsu Khola Small HP": (27.112838, 86.737766, "4P7Q+44J", "Rajapani 56200"),
    "Siddhi Khola HPP": (26.833613, 88.157672, "R5M5+C3V", "Jirmale 57300"),
    "Singati": (27.743687, 86.159188, "P5V5+FM", "Singati 45500"),
    "Super Dordi Kha": (28.271438, 84.532891, "7GCM+H5C", "Tanje 33600"),
    "Suri Khola": (27.753737, 86.219297, "Q639+FPV", "Suri 45500"),
    "Tadikhola HP": (27.922112, 85.324984, "W8CF+RXX", "F181, Ralukadevi 44900"),
    "Taksar Pikhuwa": (27.120338, 87.066016, "43C8+4CJ", "Bhojpur 57000"),   # shared point
    "Upper Chaku A": (27.873913, 85.926453, "VWFG+HH8", "Phulpingkatti"),
    "Upper Chameliya": (29.742037, 80.781828, "PQRJ+RP8", "Tapoban 10100"),
    "Upper Chirkhuwa Khola": (27.370188, 87.106062, "94C4+3C", "Shadananda 57000"),
    "Upper Gaddi Gad": (29.280938, 81.058937, "73J5+9H", "Sanagaun 10800"),
    "Upper Hugdi": (28.102613, 83.415328, "4C38+24W", "Way to Chorkate, Shantipur 32600"),
    "Upper Kalangad": (29.597538, 80.884766, "HVXM+2W6", "Banjh 10500"),
    "Upper Khorunga HPP": (27.159987, 87.544641, "5G5V+XVR", "Oyakjung 57500"),   # shared point
    "Upper Naugad Gad ": (29.723513, 80.663359, "PMF7+C85", "10100"),
    "Upper Phawa HPP": (27.318462, 87.765641, "8Q98+97J", "4, 57500"),
    "Upper Piluwa Khola 2": (27.296788, 87.385828, "79WP+P88", "Chainpur 56900"),
    "Upper Rahughat": (28.376212, 83.572359, "9HGC+FWQ", "F042, Rakhu Piple 33200"),
    "Upper Richet Khola SHP": (28.000013, 84.316672, "2828+2M3", "Yarsa 44900"),
    "Upper Sanjen": (28.183587, 85.303391, "58M3+C9J", "Chilime 45000"),
    "Upper Suri": (27.738437, 86.244062, "P6QV+9J", "Chankhu 45500"),
    "Upper Syange Khola Small": (28.386212, 84.396203, "99PW+FFP", "33600"),
    "Upper Tadi": (27.973012, 85.436922, "XCFP+6Q3", "Ghyangphedi 44900"),
    "Yambaling Khola": (27.951038, 85.790641, "XQ2R+C76", ""),
}


def _fix_districts(m):
    """Apply DISTRICT_FIX in place; returns how many rows moved."""
    if "PlantName" not in m.columns: return 0
    key = m.PlantName.astype(str).str.strip()
    n = 0
    for name, dist in DISTRICT_FIX.items():
        hit = key == name
        if hit.any() and (m.loc[hit, "DistrictId"] != dist).any():
            m.loc[hit, "DistrictId"] = dist
            n += int(hit.sum())
    return n


def resolve_coords():
    """Assign each plant a position: an exact name match where one exists,
    otherwise its district centroid, flagged approximate.

    The register's own Latitude/Longitude are ignored - every populated row holds
    39.7817,-89.6501 (Springfield, Illinois), an unedited form default. So is
    'Province 1', carried by 97 plants including many demonstrably elsewhere."""
    m = pd.read_csv(P("rms_plants_meta.csv"), dtype=str).replace(r"^\s*$", np.nan, regex=True)
    moved = _fix_districts(m)
    if moved: print(f"  district corrected on {moved} plant(s) filed under their office")
    m["PlantId"] = m.Id.astype(int)
    m["Cap_kW"] = pd.to_numeric(m.PlantCapacity, errors="coerce")
    if "Rivers" not in m.columns: m["Rivers"] = None
    cands = json.load(open(P("geo_candidates.json"), encoding="utf-8"))
    dists = json.load(open(P("np_districts.json"), encoding="utf-8"))
    ridx  = _river_index()
    taken = {}                       # river name -> points already given to a plant
    LAT, LON = (26.0, 30.7), (79.9, 88.4)

    rows = []
    for t in m.itertuples():
        hits = [c for c in cands if _is_match(t.PlantName, c["name"])
                and LAT[0] <= c["lat"] <= LAT[1] and LON[0] <= c["lon"] <= LON[1]]
        rec = {"PlantId": t.PlantId, "PlantName": t.PlantName, "Cap_kW": t.Cap_kW,
               "District": t.DistrictId, "Province": t.ProvinceId}
        pf = PLUS_FIX.get(t.PlantName)
        if pf:
            # a person checked this one against imagery; nothing here beats that
            rec |= {"lat": pf[0], "lon": pf[1], "precision": "plus",
                    "source": "Google Maps " + pf[2] + (", " + pf[3] if pf[3] else ""),
                    "matched_to": None}
        elif hits:
            hits.sort(key=lambda c: c["src"] != "OSM")
            rec |= {"lat": hits[0]["lat"], "lon": hits[0]["lon"], "precision": "exact",
                    "source": hits[0]["src"], "matched_to": hits[0]["name"]}
        elif isinstance(t.DistrictId, str) and t.DistrictId in dists:
            dd = dists[t.DistrictId]
            snap = _snap_to_river(t.PlantName, getattr(t, "Rivers", None),
                                  dd["lat"], dd["lon"], ridx, taken=taken)
            if snap:
                rec |= {"lat": snap[0], "lon": snap[1], "precision": "river",
                        "source": "on " + snap[2] + ", nearest " + t.DistrictId,
                        "matched_to": snap[2]}
            else:
                rec |= {"lat": dd["lat"], "lon": dd["lon"], "precision": "district",
                        "source": "district centroid", "matched_to": t.DistrictId}
        else:
            rec |= {"lat": None, "lon": None, "precision": "none",
                    "source": None, "matched_to": None}
        rows.append(rec)
    df = pd.DataFrame(rows)

    # a candidate claimed by two plants identifies neither (Seti HP / Seti Khola
    # HPP / Seti Nadi hpp all reduce to the same core) - demote the whole clash
    ex = df[df.precision == "exact"]
    clash = set(ex.matched_to.value_counts().pipe(lambda v: v[v > 1]).index)
    for i, row in df.iterrows():
        if row.precision == "exact" and row.matched_to in clash:
            dd = dists.get(row.District) if isinstance(row.District, str) else None
            if dd:
                sn = _snap_to_river(row.PlantName, None, dd["lat"], dd["lon"], ridx, taken=taken)
                df.loc[i, ["lat","lon","precision","source","matched_to"]] = (
                    [sn[0], sn[1], "river", "on " + sn[2] + ", nearest " + str(row.District), sn[2]]
                    if sn else [dd["lat"], dd["lon"], "district", "district centroid", row.District])
            else:
                df.loc[i, ["lat","lon","precision","source","matched_to"]] = \
                    [None, None, "none", None, None]
    df.to_csv(P("rms_plants_coords.csv"), index=False)
    print("positions:", df.precision.value_counts().to_dict())


# The register writes a company as name plus its registered office --
# "Upper Tamakoshi Hydropower Limited Gyaneshwor". The office is not the plant's
# location and is not part of the company's name, so it is cut for display. The
# CompanyId string itself is left untouched; it is the join key everywhere else.
# Whitespace is required before the place name, so a company that opens with one
# -- Butwal Power Company, Ridi Hydropower -- keeps it. Stripping from position
# zero would erase the whole name; company_name's fallback catches the rest.
CO_ADDR = re.compile(r"[,;]?\s+(?:P\.?O\.?\s*Box|Kathmandu|Katmandu|Kathmadnu|Baneshwor|"
    r"Banehswor|Newbaneswor|New\s+Baneshwor|Koteshwor|Maharajgunj|Lalitpur|Pokhara|Pokhra|"
    r"Bijaypur|Ridi|Butwal|Naxal|Thapathali|Anamnagar|Sanepa|Kupandole|Kupondole|Dillibazar|"
    r"Putalisadak|Tripureshwor|Jawalakhel|Bhaktapur|Birgunj|Biratnagar|Hetauda|Damak|Itahari|"
    r"Dharan|Nepalgunj|Galkot|Samakhusi|Sinamangal|Gongabu|Dhobighat|Tokha|Tinkune|Kamaladi|"
    r"Hativan|Pulchok|Pulchowk|Lakeside|Gyaneshwor?|Dhumbarahi?|Sunrise|Trade\s+Tower|"
    r"Tusalmarga|Bizz\s+Park|Sanothimi|Chabahil|Battisputali|Harihar|Buddhanagar|Rabibhawan|"
    r"Minbhawan|Mahalaxmisthan|Baluwatar|Thulo\s+Bharyang|Sano\s+Bharyang|Tinthana)\b.*$", re.I)


def company_name(s):
    """The company, with the registered office cut off the end.

    Falls back to the original whenever the cut leaves nothing useful. A place
    name inside a company name is not an address, and losing the company is the
    worse error of the two.
    """
    if not isinstance(s, str): return None
    full = re.sub(r"\s{2,}", " ", s).strip(" ,;-	") or None
    out = re.sub(r"[-,;\s]+$", "", re.sub(r"\s{2,}", " ", CO_ADDR.sub("", s).strip(" ,;-	")))
    return out if len(out) >= 4 else full


def plf_refs():
    """plant id -> the load factor that plant was built or contracted to reach.

    Two different claims, kept apart. A contracted PLF comes from an ICRA or CARE
    rationale and is what the PPA obliges NEA to take; a design PLF is derived
    from the mean annual energy the operator publishes. They land close, but a
    plant is judged against whichever it actually has, and a plant with neither
    is reported as such rather than measured against a fleet average.

    Two files, two granularities. scalper_plf_input.csv is hand-entered, keyed
    by ticker, and scalper_company_map.csv turns a ticker into the plants it
    runs -- so every plant a company owns gets the same figure, which is wrong
    for any company running more than one at genuinely different load factors.
    scalper_plf_projects.csv is keyed directly to a PlantId and, where a plant
    appears in both, wins: it is strictly more specific, never less trustworthy.
    """
    try:
        inp = pd.read_csv(P("scalper_plf_input.csv"), comment="#", dtype=str).fillna("")
        cmap = pd.read_csv(P("scalper_company_map.csv")).fillna("")
    except FileNotFoundError:
        return {}
    plants = {t.Ticker: [int(x) for x in str(t.PlantIds).split(";") if x]
              for t in cmap.itertuples() if t.PlantIds}
    out = {}
    for t in inp.itertuples():
        c = float(t.ContractPLF)/100 if getattr(t, "ContractPLF", "") else None
        d = float(t.DesignPLF)/100 if getattr(t, "DesignPLF", "") else None
        cost = float(t.CostPerMW) if getattr(t, "CostPerMW", "") else None
        # the annual energy the PLF was computed from, GWh -- the PLF is a rate,
        # this is the scale behind it, and the two read very differently on a
        # 3 MB run-of-river scheme versus a 300 MW one
        energy = float(t.ContractEnergy) if getattr(t, "ContractEnergy", "") else None
        if c is None and d is None and cost is None: continue
        for pid in plants.get(t.Ticker, []):
            out[pid] = {"cplf": r(c, 4), "dplf": r(d, 4), "cost_mw": r(cost, 1),
                        "cenergy": r(energy, 3),
                        # who said so: the file's own Source, except a design
                        # figure with no Source, which only ever comes from the
                        # operator's project page.
                        "psrc": (t.Source or None) or ("operator site" if d is not None else None)}

    try:
        pr = pd.read_csv(P("scalper_plf_projects.csv"), comment="#", dtype=str).fillna("")
    except FileNotFoundError:
        return out
    for t in pr.itertuples():
        c = float(t.ContractPLF)/100 if getattr(t, "ContractPLF", "") else None
        cost = float(t.CostPerMW) if getattr(t, "CostPerMW", "") else None
        energy = float(t.ContractEnergy) if getattr(t, "ContractEnergy", "") else None
        if c is None and cost is None and energy is None: continue
        pid = int(t.PlantId)
        prev = out.get(pid, {})
        # A column this row leaves blank keeps whatever the ticker-level entry had,
        # rather than blanking it -- plant-level data being more specific on PLF
        # is not a reason to erase a cost figure it happens not to carry.
        out[pid] = {"cplf": r(c, 4), "dplf": prev.get("dplf"),
                   "cost_mw": r(cost, 1) if cost is not None else prev.get("cost_mw"),
                   "cenergy": r(energy, 3) if energy is not None else prev.get("cenergy"),
                   "psrc": "project file"}
    return out


# ══════════════════════════════════════════════ listed companies -> plants ══
# The scalper series is filed per NEPSE ticker; the register is filed per plant,
# joined by a company name the register frequently misspells -- Panchakanya Mai
# "Hydropwer", Himal "Dolkha", "Surya Kund", "Snow River", "Bindhabasini",
# "Chhyandi", "Divyaswori", "Ingwa Hydopower". No normaliser reaches those, so
# scalper_company_lookup.csv holds every ticker and is read first; the matcher
# runs only for one the file has never seen.
#
# Bound at company level, never at plant level, for two reasons. The financials
# are a company's -- splitting one revenue line across the two or three plants it
# owns would be invention. And deriving the plants from the company means a plant
# joins the map on the run after it joins the register, with nothing to edit.
CO_DROP = {"ltd","limited","pvt","private","company","co","the","nepal","and","public","p","group",
           "development","project","corporation","enterprises"}
CO_EXPAND = {"hp":"hydropower","hpc":"hydropower","dev":"development","devt":"development",
             "hydro":"hydropower","power":"hydropower","hydel":"hydropower","energy":"hydropower",
             "electric":"hydropower","electricity":"hydropower","pariyojana":"project",
             "corp":"corporation","ent":"enterprises","jal":"","khola":"","nadi":"","rivers":"river"}
CO_JV = re.compile(r"^(jala?)?[bv]i?dh?ya?ut$")   # jalvidyut/jalbidhyut/vidhyut/jalavidyut
# A qualifier is the whole difference between two companies, so these are compared
# separately and must agree exactly. Super Khudi is not Khudi, Super Mai is not Mai
# Khola, Solu is not Mid Solu -- and each of those pairs really is two companies
# with two sets of filings. Same rule as _is_match applies to plant names.
#
# Words only. A bare number in a company name is part of its address -- Pokhra-6,
# Baneshwor-10, Dhobighat-3 -- never an ordinal, and reading one as a qualifier
# splits a company from itself. Roman ordinals and single letters go the same way
# for the same reason. Plant names are the opposite and QUALIFIERS keeps them.
CO_QUAL = {"upper","lower","middle","mid","madhya","mathillo","tallo","super","sano","thulo",
           "chhoto","small","mini","micro","big","main","cascade","beni"}


def _co_key(s):
    """Company name -> (core tokens, qualifier tokens).

    The register glues a postal address onto most company names -- "Union
    Hydropower Limited Dhobighat-3, Lalitpur" -- and there is no list of Kathmandu
    suburbs worth maintaining. So nothing is stripped by name; the extra tokens are
    simply tolerated by the subset test below, which is what an address is: tokens
    the listed name does not have.
    """
    if not isinstance(s, str): return frozenset(), frozenset()
    s = re.sub(r"[^a-z0-9\s]", " ", s.lower().replace(".", " ").replace("-", " ")
                                     .replace(",", " ").replace("/", " "))
    core, qual = set(), set()
    for t in s.split():
        if t.isdigit() or len(t) == 1: continue        # address fragment, not an ordinal
        if CO_JV.match(t): t = "hydropower"
        t = CO_EXPAND.get(t, t)
        if not t or t in CO_DROP: continue
        if t in CO_QUAL: qual.add(t); continue
        for a, b in TRANSLIT: t = t.replace(a, b)
        core.add(re.sub(r"(.)\1+", r"\1", t))
    return frozenset(core), frozenset(qual)


def _co_match(listed, reg):
    """The register company for `listed`, or None. reg maps key -> [names].

    Equal cores bind. Otherwise the listed name may be a subset of the register's,
    which is the address case, but only when it is a subset of exactly one -- two
    candidates mean the name does not identify a company and a human must say.
    Qualifiers must agree throughout; that is what keeps the siblings apart.
    """
    lc, lq = listed
    if not lc: return None
    hits = [names[0] for (rc, rq), names in reg.items() if rq == lq and rc == lc]
    if hits: return hits[0]
    sub = [names[0] for (rc, rq), names in reg.items()
           if rq == lq and lc < rc]
    return sub[0] if len(sub) == 1 else None


def cmd_companies(path="scalper_company_map.csv"):
    """Rebuild the listed-company -> plant map. Generated file; safe to overwrite."""
    try:
        v = pd.read_csv(P("scalper_hydro_comparative.csv"))
    except FileNotFoundError:
        print("no scalper_hydro_comparative.csv - the sync has not run"); return None
    m = pd.read_csv(P("rms_plants_meta.csv"), dtype=str).replace(r"^\s*$", np.nan, regex=True)
    m["Cap_kW"] = pd.to_numeric(m.PlantCapacity, errors="coerce")
    try:
        o = pd.read_csv(P("scalper_company_lookup.csv"), comment="#", dtype=str).fillna("")
        ov = {x.Ticker: x for x in o.itertuples()}
    except FileNotFoundError:
        ov = {}

    reg = {}
    for cid in m.CompanyId.dropna().unique():
        reg.setdefault(_co_key(cid), []).append(cid)

    rows, unmapped = [], []
    for t in v.sort_values("Ticker").itertuples():
        note = ""
        # The lookup is the authority. Matching only ever runs for a ticker it has
        # never seen, which in practice means a company that listed since the last
        # edit -- everything settled stays settled, and a register rename cannot
        # quietly rebind a company that someone already decided.
        hit = ov.get(t.Ticker)
        if hit is not None:
            regname, note = hit.RegisterCompany, hit.Note
            src = "lookup" if regname else hit.Status
        else:
            hit = None
            regname = _co_match(_co_key(t.Company), reg)
            src = "auto" if regname else "unmapped"
            if not regname: unmapped.append((t.Ticker, t.Company))
        p = m[m.CompanyId == regname] if regname else m.iloc[0:0]
        # A company the register does not carry still runs something, and the lookup
        # is the only place that is written down. Carry it through rather than
        # emitting an empty row: a project under construction is a fact about the
        # ticker, not an absence of one.
        if len(p):
            mw = r(p.Cap_kW.sum()/1000, 2)
            names = " | ".join(x.strip() for x in p.PlantName.fillna(""))
            dist = " | ".join(sorted({x for x in p.DistrictId.dropna()}))
        else:
            mw = r(hit.MW) if hit is not None else ""
            names = hit.Project if hit is not None else ""
            dist = hit.District if hit is not None else ""
        rows.append({"Ticker": t.Ticker, "Company": t.Company, "RegisterCompany": regname or "",
                     "PlantIds": ";".join(p.Id.astype(str)), "Plants": len(p),
                     "MW": mw, "PlantNames": names, "District": dist,
                     "Source": src, "Note": note})

    out = pd.DataFrame(rows)[["Ticker","Company","RegisterCompany","PlantIds","Plants",
                              "MW","PlantNames","District","Source","Note"]]
    with io.open(P(path), "w", encoding="utf-8", newline="\n") as fh:
        out.to_csv(fh, index=False)
    linked = out[out.Plants > 0]
    off = out[(out.Plants == 0) & (out.PlantNames != "")]
    print(f"companies: {len(out)} tickers {out.Source.value_counts().to_dict()}")
    print(f"  linked {len(linked)} -> {int(linked.Plants.sum())} plants, "
          f"{pd.to_numeric(linked.MW, errors='coerce').sum():.0f} MW")
    if len(off):
        print(f"  outside the register {len(off)} -> "
              f"{pd.to_numeric(off.MW, errors='coerce').sum():.0f} MW: "
              + ", ".join(f"{x.Ticker} {x.PlantNames}" for x in off.itertuples()))
    # A new listing is the routine reason for this and the one thing here that
    # needs a person. Name it, rather than letting the map sit at 109 of 110.
    if unmapped:
        print(f"  UNMAPPED {len(unmapped)}:")
        for tk, co in unmapped: print(f"    {tk:8} {co}")
        print("  add each to scalper_company_lookup.csv, blank if it files no royalties")
    return out


# ═════════════════════════════════════════════════════════════════ payload ══
def load():
    d = pd.read_csv(P("rms_monthly_clean.csv"))
    s = pd.read_csv(P("rms_summary_clean.csv"))
    m = pd.read_csv(P("rms_plants_meta.csv"), dtype=str).replace(r"^\s*$", np.nan, regex=True)
    _fix_districts(m)
    m["PlantId"]   = m.Id.astype(int)
    m["PlantName"] = m.PlantName.str.replace(r"\s+", " ", regex=True).str.strip()
    m["CodBsYear"] = m.MitiofOperation.map(bs_year)
    m["CodAdYear"] = pd.to_datetime(m.DateofOperation, errors="coerce", format="mixed").dt.year
    # format="mixed" needs pandas 2.0; on anything older it is read as a literal
    # strptime pattern, every row coerces to NaT and the page loses every AD
    # commissioning year without a word. requirements.txt pins 2.2, so this only
    # fires on a mis-set environment -- but the commit step now judges the built
    # page, which means a silent hole like that would be committed rather than
    # noticed. A column that is 198/198 populated cannot parse to nothing.
    if m.DateofOperation.notna().any() and not m.CodAdYear.notna().any():
        raise SystemExit("DateofOperation parsed to nothing - pandas is older than 2.0; "
                         "pip install -r requirements.txt")
    m["Cap_kW"]    = pd.to_numeric(m.PlantCapacity, errors="coerce")
    try: c = pd.read_csv(P("rms_plants_coords.csv"))
    except FileNotFoundError:
        c = pd.DataFrame(columns=["PlantId","lat","lon","precision","source","District"])
    return d, s, m, c


def build_payload():
    d, s, m, c = load()
    mo = d[~d.IsAnnualFiling]
    a = (d.groupby(["PlantId","PlantName","FiscalYear"], as_index=False)
           .agg(gen=("Generation_kWh","sum"), rev=("Revenue_NPR","sum"),
                roy=("Royalty_NPR","sum"), cap=("Capacity_kW","first"),
                months=("Period","size"), is_annual=("IsAnnualFiling","max")))

    # ── seasonality
    tot = mo.Generation_kWh.sum()
    season = []
    for i, name in enumerate(NEP_MONTHS, start=1):
        g = mo[mo.BsMonth == i]
        season.append({"m": i, "name": name, "greg": GREG[i-1],
                       "cf": r(g.CapacityFactor.median(), 3),
                       "cf_p25": r(g.CapacityFactor.quantile(.25), 3),
                       "cf_p75": r(g.CapacityFactor.quantile(.75), 3),
                       "gwh": r(g.Generation_kWh.sum()/1e6, 0),
                       "share": r(100*g.Generation_kWh.sum()/tot, 1),
                       "rate": r(g.Rate_NPR_kWh.median(), 2),
                       "n": int(g.CapacityFactor.notna().sum())})

    # ── fiscal years
    order = sorted(d.FiscalYear.unique(), key=FY)
    sy = s.groupby("FiscalYear").agg(due=("RoyaltyDue","sum"), recv=("Received","sum"),
                                     bal=("Balance","sum"), plants=("PlantName","nunique"))
    fy = []
    for y in order:
        g = a[a.FiscalYear == y]
        row = {"fy": y, "gwh": r(g.gen.sum()/1e6, 0),
               "plants": int(g[g.gen > 0].PlantName.nunique()),
               "mw": r(g.cap.sum()/1000, 0), "rev_bn": r(g.rev.sum()/1e9, 2)}
        if y in sy.index:
            row |= {"due_bn": r(sy.loc[y,"due"]/1e9, 2), "recv_bn": r(sy.loc[y,"recv"]/1e9, 2),
                    "bal_bn": r(sy.loc[y,"bal"]/1e9, 2), "sum_plants": int(sy.loc[y,"plants"])}
        fy.append(row)

    # ── the 15-year threshold
    sc = s[(s.Capacity_Royalty > 0) & (s.Capacity_kW > 0)].copy()
    sc["per_kw"] = sc.Capacity_Royalty/sc.Capacity_kW
    sc2 = sc.merge(m[["PlantId","CodBsYear","CodAdYear","PlantType","MitiofOperation"]],
                   on="PlantId", how="left")
    sc2["tier"] = np.where(sc2.per_kw > 700, "HIGH", "LOW")
    sc2["age"] = sc2.FiscalYear.map(FY) - sc2.CodBsYear
    ag = sc2.dropna(subset=["age"]); ag = ag[(ag.age >= 0) & (ag.age <= 45)]
    age_tier = [{"age": int(k), "low": int((g.tier=="LOW").sum()), "high": int((g.tier=="HIGH").sum()),
                 "pct": r(100*(g.tier=="HIGH").mean(), 1), "n": int(len(g))}
                for k, g in ag.groupby(ag.age.astype(int))]
    trans = []
    for name, g in sc2.sort_values("FiscalYear", key=lambda col: col.map(FY)).groupby("PlantName"):
        t, fys = g.tier.tolist(), g.FiscalYear.tolist()
        for i in range(1, len(t)):
            if t[i] == "HIGH" and t[i-1] == "LOW":
                cod = g.CodBsYear.iloc[0]
                trans.append({"plant": name, "at": fys[i], "cod": g.MitiofOperation.iloc[0],
                              "cod_ad": None if pd.isna(g.CodAdYear.iloc[0]) else int(g.CodAdYear.iloc[0]),
                              "age": None if pd.isna(cod) else int(FY(fys[i]) - cod),
                              "kw": r(g.Capacity_kW.iloc[0], 0), "type": g.PlantType.iloc[0]})
    trans.sort(key=lambda x: (x["age"] if x["age"] is not None else 99))
    lo15, hi15 = ag[ag.age < 15], ag[ag.age >= 15]
    regime = {"age_tier": age_tier, "transitions": trans,
              "n_exact15": sum(1 for t in trans if t["age"] == 15), "n_trans": len(trans),
              "under15_high_pct": r(100*lo15.tier.eq("HIGH").mean(), 1), "under15_n": int(len(lo15)),
              "over15_high_pct": r(100*hi15.tier.eq("HIGH").mean(), 1), "over15_n": int(len(hi15)),
              "companies": int(m.CompanyId.nunique()), "districts": int(m.DistrictId.nunique())}

    # ── revenue per MW, the yardstick that survives comparing a 4 MW run-of-river
    #    to a 456 MW storage scheme. Per year rather than lifetime, or a plant that
    #    has run since 2066 beats a better one commissioned last year on age alone.
    #    Only whole filing years count: a plant-year of four months is not a year.
    ry = a[(a.months >= 11) & (a.cap > 0) & (a.rev > 0) & (a.gen > 0)].copy()
    ry["rpm"] = ry.rev / (ry.cap/1000)
    ry["implied"] = ry.rev / ry.gen
    # A plant-year whose revenue over generation lands outside a plausible tariff
    # is not a cheap year or a dear one, it is a broken revenue line, and it must
    # not enter the median. Both ends matter: Khimti reconstructs at 0.54 because
    # its USD invoices go unread, and Kulekhani-II at 53.81 because one year
    # carries a settlement rather than a month of energy. Judged per plant-year,
    # not per plant -- Upper Bhotekoshi is sound in twelve years and wrong in one.
    # Judge reliability on everything the plant filed, then take the median from
    # the years that survive. Filtering first would leave nothing out of band to
    # notice, and the flag below would silently never fire again.
    implied = ry.groupby("PlantId").implied.median()
    thin = set(implied[~implied.between(2.5, 15.0)].index)
    ry = ry[ry.implied.between(2.5, 15.0)]
    rpm  = ry.groupby("PlantId").rpm.median()
    rpm_n = ry.groupby("PlantId").rpm.size()
    # Same figure, ranked once across the fleet's medians rather than year by
    # year -- "median NPR 28.9m/MW, 84th percentile of 178 plants" is a claim
    # about the plant as a whole, not about any one of its years.
    rpm_pct = rpm.rank(pct=True) * 100
    # Revenue over generation has to land near the filed PPA rate. Where it comes
    # out at a fraction of it the revenue line is short, not the tariff -- Khimti I
    # files 8.53 NPR/kWh and reconstructs to 0.54, because its invoices are in USD
    # in a shape the cleaner does not read. Flagged rather than dropped: the
    # generation is still good, and a silent hole is worse than a marked one.

    # Per-year rpm, and its percentile against every other plant's SAME fiscal
    # year -- both from this identical filtered population, so a short filing or
    # an implausible tariff can neither inflate a plant's own bar chart nor
    # contaminate the peers it gets compared against. pandas' pct rank already
    # averages ties, so two plants at an identical figure land on the same
    # percentile rather than an arbitrary coin flip deciding who ranks above whom.
    ry = ry.copy()
    ry["pct_rank"] = ry.groupby("FiscalYear").rpm.rank(pct=True) * 100
    yr_n = ry.groupby("FiscalYear").rpm.size()
    rpm_yearly = {(int(t.PlantId), t.FiscalYear):
                  (r(t.rpm/1e6, 2), r(t.pct_rank, 1), int(yr_n[t.FiscalYear]))
                  for t in ry.itertuples()}


    # ── plant table (PPA rates split wet/dry - the tariff is seasonal)
    wet = mo[~mo.BsMonth.isin(DRY_MONTHS)].groupby("PlantId").Rate_NPR_kWh.median()
    dry = mo[mo.BsMonth.isin(DRY_MONTHS)].groupby("PlantId").Rate_NPR_kWh.median()
    allr = mo.groupby("PlantId").Rate_NPR_kWh.median()
    usd_plants = (d.groupby("PlantId").Revenue_USD.count() > 0).to_dict()
    usd_rate = d.groupby("PlantId").Rate_USD_kWh.median()
    usd_fx   = d.groupby("PlantId").FX.median()
    pl = (a.groupby("PlantName", as_index=False)
            .agg(gwh=("gen", lambda x: x.sum()/1e6), rev=("rev", lambda x: x.sum()/1e9),
                 roy=("roy", lambda x: x.sum()/1e9), kw=("cap","first"),
                 yrs=("FiscalYear","nunique"), pid=("PlantId","first")))
    pl = pl.merge(mo.groupby("PlantName").CapacityFactor.median().rename("cf"),
                  on="PlantName", how="left").sort_values("gwh", ascending=False)
    s2 = s.copy(); s2["_o"] = s2.FiscalYear.map(FY)
    last = s2.sort_values("_o").groupby("PlantName").tail(1).set_index("PlantName")
    refs = plf_refs()
    meta = m.set_index("PlantId")
    coord = c.set_index("PlantId") if len(c) else None
    g_ = lambda row, k: None if k not in row.index or pd.isna(row[k]) else str(row[k])
    plants = []
    for x in pl.itertuples():
        pid = int(x.pid)
        rec = {"id": pid, "name": x.PlantName, "mw": r(x.kw/1000, 1), "gwh": r(x.gwh, 0),
               "cf": r(x.cf, 3), "rev_bn": r(x.rev, 2), "roy_bn": r(x.roy, 3),
               "yrs": int(x.yrs), "bal_m": r(last.Balance.get(x.PlantName, np.nan)/1e6, 1),
               "last_fy": last.FiscalYear.get(x.PlantName),
               "ppa": r(allr.get(pid), 2), "ppa_wet": r(wet.get(pid), 2), "ppa_dry": r(dry.get(pid), 2),
               "rpm": r(rpm.get(pid, np.nan)/1e6, 1), "rpm_n": int(rpm_n.get(pid, 0)),
               "rpm_pct": r(rpm_pct.get(pid), 1),
               "cplf": refs.get(pid, {}).get("cplf"), "dplf": refs.get(pid, {}).get("dplf"),
               "cost_mw": refs.get(pid, {}).get("cost_mw"), "psrc": refs.get(pid, {}).get("psrc"),
               "cenergy": refs.get(pid, {}).get("cenergy"),
               "implied": r(implied.get(pid, np.nan), 2), "thin": 1 if pid in thin else 0,
               "ccy": "USD" if usd_plants.get(pid) else "NPR",
               "ppa_usd": r(usd_rate.get(pid), 4), "fx": r(usd_fx.get(pid), 2)}
        if pid in meta.index:
            mr = meta.loc[pid]
            rec |= {"company": (company_name(g_(mr,"CompanyId")) or "")[:60] or None,
                    "district": g_(mr,"DistrictId"), "province": g_(mr,"ProvinceId"),
                    "cod": g_(mr,"MitiofOperation"),
                    "cod_ad": None if pd.isna(mr.CodAdYear) else int(mr.CodAdYear),
                    "ptype": g_(mr,"PlantType"), "lic": g_(mr,"LicenseNumber"),
                    "lic_from": g_(mr,"LicensedMiti"), "lic_to": g_(mr,"LicenseExpiryMiti"),
                    "river": g_(mr,"Rivers")}
        if coord is not None and pid in coord.index:
            cr = coord.loc[pid]
            rec |= {"prec": g_(cr,"precision"), "lat": r(cr.lat, 4), "lon": r(cr.lon, 4)}
        plants.append(rec)

    # ── per plant-year detail
    moy = (d.groupby(["PlantId","FiscalYear"], as_index=False)
             .agg(gen=("Generation_kWh","sum"), rev=("Revenue_NPR","sum"),
                  roy_m=("Royalty_NPR","sum"), rate=("Rate_NPR_kWh","median"),
                  months=("Period","size"), annual=("IsAnnualFiling","max")))
    su = s.groupby(["PlantId","FiscalYear"], as_index=False).agg(
            eroy=("Energy_Royalty","sum"), croy=("Capacity_Royalty","sum"),
            due=("RoyaltyDue","sum"), recv=("Received","sum"), bal=("Balance","sum"))
    x = moy.merge(su, on=["PlantId","FiscalYear"], how="outer") \
           .merge(m[["PlantId","CodBsYear","Cap_kW"]], on="PlantId", how="left")
    x["age"] = x.FiscalYear.map(FY) - x.CodBsYear
    fin = lambda v: v is not None and np.isfinite(v)
    years = {}
    for pid, g in x.groupby("PlantId"):
        rows = []
        for t in g.sort_values("FiscalYear", key=lambda col: col.map(FY)).itertuples():
            eroy = t.eroy if fin(t.eroy) else (t.roy_m if fin(t.roy_m) else np.nan)
            due = t.due if fin(t.due) else eroy
            pct = 100*eroy/t.rev if (fin(t.rev) and t.rev > 1e6 and fin(eroy)) else np.nan
            # capacity royalty of zero means "not assessed yet", not "tier 1"
            tier = None
            if fin(t.croy) and t.croy > 0 and fin(t.Cap_kW) and t.Cap_kW > 0:
                tier = 2 if (t.croy/t.Cap_kW) > 700 else 1
            # None here means this specific plant-year didn't clear the same two
            # gates the lifetime rpm median applies (a whole filing, a plausible
            # implied tariff) -- not that the revenue or capacity is missing.
            yrpm, ypct, yn = rpm_yearly.get((int(pid), t.FiscalYear), (None, None, None))
            rows.append([t.FiscalYear, r(t.gen/1e6,1) if fin(t.gen) else None,
                         r(t.rev/1e6,1) if fin(t.rev) else None,
                         r(eroy/1e6,2) if fin(eroy) else None,
                         r(t.croy/1e6,2) if fin(t.croy) else None,
                         r(due/1e6,2) if fin(due) else None,
                         r(t.recv/1e6,2) if fin(t.recv) else None,
                         r(t.bal/1e6,2) if fin(t.bal) else None,
                         r(pct,2) if fin(pct) else None, tier,
                         int(t.age) if fin(t.age) else None,
                         int(t.months) if fin(t.months) else 0,
                         1 if (fin(t.annual) and t.annual) else 0,
                         r(t.rate, 2) if fin(t.rate) else None,
                         yrpm, ypct, yn])
        years[str(int(pid))] = rows

    # ── monthly detail (third drill-down level)
    months = {}
    mcols = ["BsYear","BsMonth","Generation_kWh","Revenue_NPR","Royalty_NPR",
             "Rate_NPR_kWh","CapacityFactor","Revenue_USD","Rate_USD_kWh","FX"]
    for (pid, fyy), g in d.sort_values(["BsYear","BsMonth"]).groupby(["PlantId","FiscalYear"]):
        months.setdefault(str(int(pid)), {})[fyy] = [
            [int(t.BsYear)-2000, int(t.BsMonth), r(t.Generation_kWh/1e6, 2),
             r(t.Revenue_NPR/1e6, 2), r(t.Royalty_NPR/1e6, 3),
             r(t.Rate_NPR_kWh, 2), r(t.CapacityFactor, 3),
             r(t.Revenue_USD/1e6, 3), r(t.Rate_USD_kWh, 4), r(t.FX, 2)]
            for t in g[mcols].itertuples(index=False)]

    # ── licensing pipeline: when licences were granted, and how long each took
    #    to reach commercial operation
    lic_y, cod_y = m.LicensedMiti.map(bs_year), m.CodBsYear
    exp_y = m.LicenseExpiryMiti.map(bs_year)
    lead  = (cod_y - lic_y).where(lambda v: v.between(-2, 25))
    yrs   = sorted(set(int(v) for v in pd.concat([lic_y, cod_y]).dropna()
                       if 2040 <= v <= latest_bs(d)))
    lic_rows = []
    for y in yrs:
        ln = int((lic_y == y).sum()); cn = int((cod_y == y).sum())
        sub = lead[(lic_y == y) & lead.notna()]
        lic_rows.append({"y": y, "lic": ln, "cod": cn,
                         "lead": r(sub.median(), 1) if len(sub) >= 3 else None,
                         "n": int(len(sub))})
    licence = {"byYear": lic_rows,
               "median_lead": r(lead.median(), 1),
               "p25_lead": r(lead.quantile(.25), 1), "p75_lead": r(lead.quantile(.75), 1),
               "n_lead": int(lead.notna().sum()),
               "term_median": r((exp_y - lic_y).median(), 0),
               "peak_year": int(lic_y.value_counts().idxmax()),
               "peak_count": int(lic_y.value_counts().max()),
               "first_year": yrs[0] if yrs else None, "last_year": yrs[-1] if yrs else None}

    payload = {"season": season, "fy": fy, "plants": plants,
               "licence": licence,
               "regime": regime, "years": years, "months": months,
               "nepMonths": NEP_MONTHS, "greg": GREG}

    mp = _map_payload(d, m, c, plants)
    if mp: payload["map"] = mp
    rc = _recon_payload(d, m)
    if rc: payload["recon"] = rc
    co = _co_payload(d, m)
    if co: payload["co"] = co

    payload["stats"] = {
        "plants": int(d.PlantName.nunique()),
        "mw": r(d.groupby("PlantName").Capacity_kW.first().sum()/1000, 0),
        "twh": r(a.gen.sum()/1e9, 1),
        "royalty_bn": r(s.RoyaltyDue.sum()/1e9, 1),
        "received_bn": r(s.Received.sum()/1e9, 1),
        "fy_span": f"{order[0]} – {order[-1]}", "n_fy": len(order),
        "latest_fy": order[-1], "rows": int(len(d)),
        "companies": int(m.CompanyId.nunique()),
        "peak_month": max(season, key=lambda q: q["cf"] or 0)["name"],
        "trough_month": min(season, key=lambda q: q["cf"] or 9)["name"],
        "plant_years": sum(len(v) for v in years.values()),
        "month_rows": sum(len(v2) for v in months.values() for v2 in v.values()),
        "median_lead": licence["median_lead"], "licence_peak": licence["peak_year"],
        "cod_known": int(m.CodBsYear.notna().sum()), "cod_total": int(len(m)),
        # when the register was last read, taken from the raw scrape file rather
        # than from today - a rebuild does not make the data any fresher
        "retrieved": _retrieved()}
    return payload


def _chain(segs):
    """OSM splits one river into consecutive ways, so a 60 km drainage arrives as
    fifteen fragments none of which is long enough to hang a label on. Stitch the
    ones that share an endpoint back into continuous polylines: labels get a run
    to sit along, and hit-testing and rendering both get cheaper for it."""
    from collections import defaultdict
    segs = [[tuple(p) for p in s] for s in segs if len(s) >= 2]
    ends = defaultdict(list)
    for i, s in enumerate(segs):
        ends[s[0]].append(i)
        ends[s[-1]].append(i)
    used, out = set(), []
    for i, s in enumerate(segs):
        if i in used: continue
        used.add(i)
        ch = list(s)
        for _ in range(2):                      # extend one way, flip, extend the other
            while True:
                nxt = next((j for j in ends.get(ch[-1], ()) if j not in used), None)
                if nxt is None: break
                used.add(nxt)
                t = segs[nxt]
                ch += t[1:] if t[0] == ch[-1] else t[-2::-1]
            ch.reverse()
        out.append([list(p) for p in ch])
    return out


def _watercourses(terr, m, c):
    """One selectable entity per real watercourse, plus the river each plant sits on.

    OSM names a single river several ways along its length -- 'Kali Gandaki' and
    'Kali Gandaki River' are one river in two rows, as are 'Seti Nadi' / 'Seti
    Khola' / 'Seti River' -- so segments are merged on the normalised core name
    and the longest variant supplies the label. Otherwise picking a river would
    mean picking which spelling of it you wanted.

    A plant is attributed in the order that keeps the map honest: the watercourse
    its marker was snapped onto first, so the dot and the claim never disagree,
    then the river the register names for it, then its own name."""
    groups = {}
    for rv in terr.get("rivers", []) + terr.get("streams", []):
        # a Devanagari name normalises to nothing; key it raw so it still draws
        core = _parts(rv["n"])[0]
        k = frozenset(core) if core else "raw:" + rv["n"]
        g = groups.get(k)
        if g is None:
            g = groups[k] = {"n": rv["n"], "km": 0.0, "best": -1.0, "segs": [], "sm": 1}
        g["km"] += rv["km"]
        g["segs"] += rv["segs"]
        if not rv.get("sm"): g["sm"] = 0          # any main-stem variant -> main stem
        if rv["km"] > g["best"]: g["best"], g["n"] = rv["km"], rv["n"]

    order = sorted(groups.items(), key=lambda kv: -kv[1]["km"])
    idx = {k: k for k, _ in order if isinstance(k, frozenset)}
    for k, _ in order:                            # single-token fallback, longest first
        if isinstance(k, frozenset):
            for t in k: idx.setdefault(frozenset([t]), k)

    def resolve(*srcs):
        for s in srcs:
            for one in (_named_rivers(s) if s is not None else []):
                core = _parts(one)[0]
                if not core: continue
                hit = idx.get(frozenset(core))
                if hit: return hit
                for t in sorted(core):            # sorted: never depend on set order
                    hit = idx.get(frozenset([t]))
                    if hit: return hit
        return None

    pos = {t.PlantId: t for t in c.itertuples()}
    rivers = {}                                   # register river name by plant id
    if "Rivers" in m.columns:
        rivers = dict(zip(m.PlantId, m.Rivers))
    gid, out = {}, []
    for k, g in order:
        gid[k] = len(out)
        out.append({"n": g["n"], "km": int(round(g["km"])), "segs": _chain(g["segs"]),
                    "sm": g["sm"], "np": 0, "mw": 0.0})

    by_plant = {}
    for pid, t in pos.items():
        snap = t.matched_to if getattr(t, "precision", None) == "river" else None
        k = resolve(snap, rivers.get(pid), t.PlantName)
        if k is None: continue
        by_plant[pid] = gid[k]
        out[gid[k]]["np"] += 1
        kw = getattr(t, "Cap_kW", None)
        if kw is not None and np.isfinite(kw): out[gid[k]]["mw"] += kw/1000.0
    for w in out: w["mw"] = r(w["mw"], 1)
    return out, by_plant


def _map_payload(d, m, c, plants):
    try: outline = json.load(open(P("np_outline.json"), encoding="utf-8"))
    except FileNotFoundError: return None
    gen = (d.groupby("PlantId").Generation_kWh.sum()/1e6).rename("gwh")
    cf  = d[~d.IsAnnualFiling].groupby("PlantId").CapacityFactor.median().rename("cf")
    x = c.merge(m[["PlantId","CompanyId","MitiofOperation","CodBsYear"]], on="PlantId", how="left") \
         .merge(gen, on="PlantId", how="left").merge(cf, on="PlantId", how="left")
    x = x[x.lat.notna()].copy()
    GOLDEN = math.pi * (3 - math.sqrt(5))
    x["jlat"], x["jlon"] = x.lat, x.lon
    for _, grp in x[x.precision == "district"].groupby("District"):
        for k, i in enumerate(grp.index):
            rad, ang = 0.055*math.sqrt(k+0.6), (k+1)*GOLDEN
            x.loc[i,"jlat"] = x.loc[i,"lat"] + rad*math.sin(ang)
            x.loc[i,"jlon"] = x.loc[i,"lon"] + rad*math.cos(ang)/math.cos(math.radians(28.4))
    # Whatever still lands on one coordinate: three of the looked-up pairs share a
    # reference point, and two plants on different rivers can meet at a confluence
    # the snap sent both to. Drawn at the same spot they are one dot that only the
    # last one drawn can be hovered or clicked, so ring them at ~300 m - well
    # inside the error these positions already carry. District centroids are
    # excluded because the loop above has already spread them much further.
    for _, grp in x[x.precision != "district"].groupby(["lat", "lon"]):
        if len(grp) < 2: continue
        for k, i in enumerate(grp.index):
            ang = k * 2*math.pi/len(grp)
            x.loc[i,"jlat"] = x.loc[i,"lat"] + 0.0027*math.sin(ang)
            x.loc[i,"jlon"] = x.loc[i,"lon"] + 0.0027*math.cos(ang)/math.cos(math.radians(28.4))
    try:
        terrain = json.load(open(P("np_terrain.json"), encoding="utf-8"))
    except FileNotFoundError:
        terrain = {"rivers": [], "streams": [], "peaks": []}
    wcs, on_river = _watercourses(terrain, m, x)
    ids = {key(p["name"]): p["id"] for p in plants}
    pts, linked = [], 0
    for t in x.itertuples():
        age = None if not np.isfinite(t.CodBsYear) else int(latest_bs(d) - t.CodBsYear)
        pid = ids.get(key(t.PlantName))
        if pid is not None: linked += 1
        pts.append({"id": pid, "n": t.PlantName,
                    "co": None if pd.isna(t.CompanyId) else str(t.CompanyId)[:52],
                    "mw": r(t.Cap_kW/1000.0, 1), "la": r(t.jlat, 4), "lo": r(t.jlon, 4),
                    "p": 2 if t.precision in ("exact", "plus") else (1 if t.precision == "river" else 0),
                    "src": None if pd.isna(t.source) else str(t.source),
                    "di": None if pd.isna(t.District) else str(t.District),
                    "cod": None if pd.isna(t.MitiofOperation) else str(t.MitiofOperation),
                    "age": age, "t2": 1 if (age is not None and age >= 15) else 0,
                    "rv": on_river.get(t.PlantId),
                    "gwh": r(t.gwh, 0), "cf": r(t.cf, 3)})
    pts.sort(key=lambda q: -(q["mw"] or 0))
    try:
        grid = json.load(open(P("np_grid.json"), encoding="utf-8"))
    except FileNotFoundError:
        grid = {"lines": [], "subs": []}
    try:
        plan = json.load(open(P("np_grid_plan.json"), encoding="utf-8"))
    except FileNotFoundError:
        plan = {"lines": [], "labels": {}}
    return {"pts": pts, "outline": outline["rings"], "bbox": outline["bbox"],
            "rivers": wcs, "n_onriver": sum(1 for q in pts if q["rv"] is not None),
            "peaks": terrain["peaks"],
            "lines": grid["lines"], "subs": grid["subs"],
            "plan": plan["lines"], "plan_labels": plan.get("labels", {}),
            "plan_subs": plan.get("subs", []),
            "plan_plants": plan.get("plants", []),
            "plan_plant_labels": plan.get("plant_labels", {}),
            "plan_sub_labels": plan.get("sub_labels", {}),
            "plan_clusters": plan.get("clusters", []),
            "plan_cluster_label": plan.get("cluster_label"),
            "plan_source": plan.get("source"),
            "n_exact": int((c.precision == "exact").sum()),
            "n_plus": int((c.precision == "plus").sum()),
            "n_river": int((c.precision == "river").sum()),
            "n_district": int((c.precision == "district").sum()),
            "n_none": int((c.precision == "none").sum()), "n_total": int(len(c)),
            "linked": linked, "mw_located": r(x.Cap_kW.sum()/1000.0, 0),
            "mw_total": r(m.Cap_kW.sum()/1000.0, 0)}


def _reported_quarters():
    """The operator's published cumulative energy revenue, {fy: {Qn: NPR}}.

    Read from the workbook when it is there, and cached to recon_reported.json
    so a build without it -- CI always, anyone who is not the author -- can still
    draw the section. Only this half is cached: it is the half that cannot be
    recomputed. The RMS half is rebuilt from live scraped data on every build,
    so a monthly refresh moves the comparison instead of freezing it.
    """
    cache = P("recon_reported.json")
    if os.path.exists(XLSX):
        try:
            import openpyxl
            ws = openpyxl.load_workbook(XLSX, data_only=True)["is"]
        except Exception:
            ws = None
        if ws is not None:
            cum = {}
            for i in range(2, ws.max_column+1):
                y, q, v = ws.cell(1,i).value, ws.cell(2,i).value, ws.cell(3,i).value
                if y and isinstance(v, (int, float)):
                    cum.setdefault(str(y), {})[str(q)] = float(v)
            if cum:
                json.dump({"plant": "Likhu-4", "plant_id": 116, "mw": 52.4,
                           "measure": "Income from Sale of Energy, cumulative year to date, NPR",
                           "source": "Green Ventures Co. Ltd published quarterly income statements",
                           "cum": cum},
                          open(cache, "w", encoding="utf-8"), indent=1, sort_keys=True)
                return cum
    try:
        return json.load(open(cache, encoding="utf-8"))["cum"]
    except (FileNotFoundError, KeyError, ValueError):
        return None


MKT_COLS = {
    # what it is here          the column as the comparative file spells it
    "px":      "Latest Close",          "avg180":  "180 Day Avg",
    "chg1w":   "1 Wk Price Chg %",      "chg4w":   "4 Wk Price Chg %",
    "chg12w":  "12 Wk Price Chg %",     "chgytd":  "YTD Price Chg %",
    "chg1y":   "1 Year Price Chg %",    "mcap":    "Market Cap",
    "paid":    "Paidup Capital",        "res":     "Reserves & Surplus",
    "rev":     "Revenue",               "ni":      "Net Income",
    "gm":      "Gross Margin TTM %",    "em_":     "EBITDA Margin TTM %",
    "npm":     "Net Profit Margin TTM %",
    "roe":     "ROE (TTM) %",           "roa":     "ROA (TTM)%",
    "eps":     "EPS (TTM)",             "bv":      "Bookvalue",
    "pe":      "PE (TTM)",              "pbv":     "PBV",
    "cr":      "Current Ratio",         "qr":      "Quick Ratio",
    "d2a":     "Debt to Asset",         "eqm":     "Equity Multiplier",
    "beta":    "Beta Weekly",           "betam":   "Beta Monthly",
    "var1m":   "VaR 1 Month @ 5%",      "sd":      "S.D of Returns [Wk, 1Y]%",
    "mret":    "Mean Returns [Wk, 1Y]%", "rsi":    "RSI",
    "rstr":    "Relative Strength %",   "ma":      "50 day MA vs 200 day MA %",
    "vs52":    "Price Vs 52 Week High %", "macd":  "MACD Crossover",
}
# money in this file is NPR thousands, except Market Cap which is NPR
MKT_SCALE = {"mcap": 1e-6, "paid": 1e-3, "res": 1e-3, "rev": 1e-3, "ni": 1e-3}

# The income statement as the cascade it is, in the order and under the names
# the filings themselves use -- not a paraphrase, so a reader checking this
# against the primary source can match every line by name. Sign says which
# way the bar points: +1 adds to the running total, -1 takes away, 0 is a
# subtotal, checked against its own filed column rather than accumulated
# and trusted.
CASCADE = [("EnergySales", +1, "Income from Sale of Energy"),
           ("CostOfProduction", -1, "Cost of Production"),
           ("GrossProfit", 0, "Gross Profit"),
           ("DividendIncome", +1, "Income from Dividend"),
           ("ForexGainLoss", +1, "Forex Gain"),
           ("OtherIncome", +1, "Income from Other Sources"),
           ("AdminExpenses", -1, "Administrative Expenses"),
           ("OperatingProfit", 0, "Operating Profit"),
           ("InterestIncomeExpense", +1, "Net Interest Income"),
           # Only about a quarter of filers break this out; the rest fold it into
           # cost of production. Where it is filed it is exactly what closes the
           # chain from operating profit to profit before tax, and leaving it out
           # put an 18m unreconciled bar on every one of those companies.
           ("Depreciation_IncomeStatement", -1, "Depreciation"),
           ("Provisions", -1, "Gross Provisions"),
           ("ProfitBeforeTax", 0, "Profit Before Taxes"),
           ("Taxes", -1, "Taxes"),
           ("Bonus", -1, "Bonuses"),
           ("ProfitAfterTax", 0, "Net Profit")]

# The balance sheet, complete rather than a summary of it -- every line the
# filing itself carries, same order, same names, so this can be checked
# line by line against the primary source. Third field marks a subtotal:
# bold in the table, and (Total Sources of Funds, Application of Funds)
# ought to equal each other by construction, the same identity a T-account
# balance sheet is built to enforce.
BS_LINES = [("PaidUpCapital", "Paid up Capital", False),
            ("Premium", "Share Premium", False),
            ("Reserves", "Reserves", False),
            ("LtLiabilities", "Long Term Liabilities", False),
            ("TotalFunds", "Total Sources of Funds", True),
            ("FixedAssets", "Fixed Assets", False),
            ("Depreciation", "Depreciation", False),
            ("NetFixedAssets", "Net Fixed Assets", False),
            ("NonCoreAssets", "Non Core Assets", False),
            ("Investments", "Investments", False),
            ("WorkInProgress", "Work in Progress", False),
            ("Cash", "Cash at Hand", False),
            ("Receivables", "Receivables", False),
            ("AdvancesPrepaymentsLoansDeposits", "Advances, Payments, Loans and Deposits", False),
            ("Inventory", "Inventory", False),
            ("TotalCurrentAssets", "Total Current Assets", True),
            ("StLiabilities", "Short Term Liabilities", False),
            ("DeferredLiabilities", "Deferred Liabilities", False),
            ("TotalStLiabilities", "Total Short Term Liabilities", True),
            ("ApplicationOfFunds", "Application of Funds", True)]


def _loo_slope_range(x, y):
    """How far one company can move a straight-line fit. A relationship a single
    row can invert is not one, and this is the check that says so."""
    x, y = np.asarray(x, float), np.asarray(y, float)
    if len(x) < 6: return None, None
    out = []
    for i in range(len(x)):
        k = np.ones(len(x), bool); k[i] = False
        out.append(np.polyfit(x[k], y[k], 1)[0])
    return float(min(out)), float(max(out))


def _fit(x, y):
    x, y = np.asarray(x, float), np.asarray(y, float)
    ok = np.isfinite(x) & np.isfinite(y)
    x, y = x[ok], y[ok]
    if len(x) < 6: return None
    sl, ic = np.polyfit(x, y, 1)
    r2 = 1 - ((y-(ic+sl*x))**2).sum()/((y-y.mean())**2).sum()
    lo, hi = _loo_slope_range(x, y)
    return {"slope": r(float(sl), 4), "icept": r(float(ic), 4), "r2": r(float(r2), 3),
            "n": int(len(x)), "loo_lo": r(lo, 3), "loo_hi": r(hi, 3),
            "flips": 1 if (lo is not None and lo*hi < 0) else 0}


def _market(rows):
    """The price screen, the quarterly statements and the two fits the value tab
    turns on. Everything here is per ticker and joins onto the rows already built.

    The banking version of this page fits price-to-book on return on equity, which
    is the textbook relationship for a lender. It does not survive contact with a
    hydro cohort: one company at 33.8x book carries the whole of an R-squared of
    0.24, and leaving it out drops that to 0.03 and inverts the slope. So no line
    is fitted to it and the scatter is shown bare, with the leave-one-out range
    printed beside it.

    What does hold is capacity. Market capitalisation on installed MW, log on log,
    explains about three fifths of the variance, and no single company moves the
    slope by more than 0.02. It is markedly sublinear -- so the market pays less
    per megawatt the bigger the plant, and the residual from that line is a far
    better reading of dear and cheap than a price-to-book gap would have been.
    """
    try:
        v = pd.read_csv(P("scalper_hydro_comparative.csv"))
    except FileNotFoundError:
        return None
    v = v.drop_duplicates("Ticker").set_index("Ticker")
    by = {x["tk"]: x for x in rows}
    for tk, x in by.items():
        if tk not in v.index: continue
        src = v.loc[tk]
        for k, col in MKT_COLS.items():
            if col not in v.columns: continue
            val = pd.to_numeric(src[col], errors="coerce")
            x[k] = None if not np.isfinite(val) else r(float(val)*MKT_SCALE.get(k, 1.0),
                                                       0 if k in ("mcap","paid","res","rev","ni") else 4)
        # NEPSE par is NPR 100. Under it the company is into its paid-up capital
        # and every ratio with equity in the denominator stops being comparable.
        x["subpar"] = 1 if (x.get("bv") is not None and x["bv"] <= 100) else 0
        # the screened cohort the market tab runs on: above par and earning
        x["scr"] = 1 if (not x["subpar"] and x.get("bv") is not None
                         and x.get("eps") is not None and x["eps"] > 0) else 0
        x["mcap_m"] = x.get("mcap")
        x["mcap_mw"] = r(x["mcap"]/x["mw"], 0) if x.get("mcap") and x.get("mw") else None
        # retained earnings against paid-up: the dividend that could be declared
        # out of what has actually accumulated. A balance, never annualised.
        x["payable"] = r(100*x["res"]/x["paid"], 1) if x.get("res") is not None and x.get("paid") else None

    keep = [x for x in rows if x.get("mcap") and x.get("mw")]
    cap = _fit([math.log10(x["mw"]) for x in keep], [math.log10(x["mcap"]) for x in keep])
    if cap:
        for x in rows:
            if x.get("mcap") and x.get("mw"):
                f = 10**(cap["icept"] + cap["slope"]*math.log10(x["mw"]))
                x["cap_fit"] = r(f, 0)
                x["cap_gap"] = r(x["mcap"]/f - 1, 3)
        cap["doubling"] = r(2**cap["slope"], 3)

    pb = [x for x in rows if x.get("pbv") is not None and x.get("roe") is not None]
    pbroe = _fit([x["roe"] for x in pb], [x["pbv"] for x in pb])
    par = [x for x in pb if not x.get("subpar")]
    pbroe_par = _fit([x["roe"] for x in par], [x["pbv"] for x in par])

    # cohort medians, so every figure on the company tab has something to sit
    # beside. The book-derived ones are struck on the above-par cohort only, for
    # the same reason the chart is: a median that includes a 33.8x and a -10.9x
    # is not describing the same quantity as the rest of the column.
    BOOK = {"pbv", "bv", "roe", "eqm"}
    KEYS = list(MKT_COLS) + ["mw","gwh","plf","rpm","mcap_mw","payable","cap_gap","ratio","ach"]
    med, smed = {}, {}
    for k in KEYS:
        src = [x for x in rows if not x.get("subpar")] if k in BOOK else rows
        vals = [x[k] for x in src if x.get(k) is not None]
        if vals: med[k] = r(float(np.median(vals)), 4)
        # and the same medians struck on the screened cohort alone, which is what
        # the market tab compares against
        svals = [x[k] for x in rows if x.get("scr") and x.get(k) is not None]
        if svals: smed[k] = r(float(np.median(svals)), 4)
    n_par = sum(1 for x in rows if x.get("bv") is not None and not x.get("subpar"))
    scr = [x for x in rows if x.get("scr")]
    return {"cap": cap, "pbroe": pbroe, "pbroe_par": pbroe_par,
            "med": med, "smed": smed,
            "n_mkt": sum(1 for x in rows if x.get("mcap")),
            "n_par": n_par, "n_subpar": sum(1 for x in rows if x.get("subpar")),
            "subpar_mcap": r(sum(x.get("mcap") or 0 for x in rows if x.get("subpar")), 0),
            "all_mcap": r(sum(x.get("mcap") or 0 for x in rows), 0),
            "n_scr": len(scr),
            "scr_mcap": r(sum(x.get("mcap") or 0 for x in scr), 0),
            "n_loss": sum(1 for x in rows if x.get("eps") is not None and x["eps"] <= 0),
            # above par but loss-making: what the earnings test removes on its own
            "n_loss_par": sum(1 for x in rows if not x.get("subpar")
                              and x.get("eps") is not None and x["eps"] <= 0),
            "book_keys": sorted(BOOK)}


def _quarters(rows):
    """Eight discrete quarters per ticker out of a cumulative filing.

    Every line in this series is year to date and resets at Q1, so printing it as
    filed would show energy sales collapsing between Q4 and Q1 -- which is the
    fiscal year restarting and nothing else. Q1 stands as filed; every other
    quarter is itself minus the one before it, and only when that one is actually
    there. A gap leaves a blank rather than a difference across it, because a Q3
    with no Q2 differenced against Q1 is two quarters of trading reported as one.
    """
    try:
        f = pd.read_csv(P("scalper_hydro_combined.csv"))
    except FileNotFoundError:
        return {}
    f["_aud"] = (f.DataSource == "Audited").astype(int)
    f = f.sort_values("_aud").drop_duplicates(["Ticker","Year","Quarter"], keep="last")
    f["fy0"] = f.Year.map(lambda s: int(str(s).split("/")[0]))
    f = f.sort_values(["fy0","Quarter"])
    cols = [c for c, _, _ in CASCADE] + [c for c, _, _ in BS_LINES]
    cols = [c for c in dict.fromkeys(cols) if c in f.columns]
    flow = {c for c, _, _ in CASCADE} | {"EpsAnnualized"}

    out = {}
    tickers = {x["tk"] for x in rows}
    for tk, g in f.groupby("Ticker"):
        if tk not in tickers: continue
        g = g.tail(9)
        prev = {}
        prev_src = {}                 # (fy0, quarter) -> Published/Audited
        last_shares = None            # forward-filled: a share count does not
        qs = []                       # change every quarter, only on an issue
        for t in g.itertuples():
            rec = {"y": t.Year, "q": int(t.Quarter)}
            # An audited year-end restatement is real and belongs in the
            # cumulative figure -- but Q1-Q3 almost never get their own audit,
            # so a restated Q4 differenced against an unaudited Q3 baseline
            # mixes two accounting bases into one "quarter". Found on GVL:
            # Q4 2021/22 audited cumulative PAT minus Q3's published baseline
            # gives a standalone quarter 74% above what the same subtraction
            # gives using the published Q4 instead (42.1m against 24.2m) --
            # not GVL's fourth quarter, mostly the audit's own correction to
            # the first three, dumped entirely into the last one by arithmetic.
            basis_ok = prev_src.get((t.fy0, t.Quarter - 1)) in (None, t.DataSource)
            for c in cols:
                cum = pd.to_numeric(getattr(t, c, np.nan), errors="coerce")
                if not np.isfinite(cum):
                    rec[c] = None; continue
                if c not in flow:                       # a balance is as filed
                    rec[c] = r(float(cum)/1e3, 2); continue
                # Year-to-date exactly as filed, alongside the discrete figure --
                # the toggle between them is a display choice, not a second
                # computation, so both come from this one pass over the row.
                rec[c + "_cum"] = r(float(cum)/1e3, 2)
                if t.Quarter == 1:
                    rec[c] = r(float(cum)/1e3, 2)
                elif not basis_ok:
                    rec[c] = None
                else:
                    p = prev.get((c, t.fy0, t.Quarter-1))
                    rec[c] = None if p is None else r((float(cum)-p)/1e3, 2)
            for c in cols:
                cum = pd.to_numeric(getattr(t, c, np.nan), errors="coerce")
                if np.isfinite(cum): prev[(c, t.fy0, int(t.Quarter))] = float(cum)
            prev_src[(t.fy0, int(t.Quarter))] = t.DataSource

            # Book value per share is a point in time, like any other balance-
            # sheet line -- shown as filed, never differenced. The filing splits
            # it across two columns depending which statement type the row came
            # from (BookValuePerShare or NetWorthPerShare); at most one is ever
            # populated on a given row, so this is a coalesce, not a choice.
            bv = pd.to_numeric(getattr(t, "BookValuePerShare", np.nan), errors="coerce")
            if not np.isfinite(bv):
                bv = pd.to_numeric(getattr(t, "NetWorthPerShare", np.nan), errors="coerce")
            if np.isfinite(bv): rec["BookValue"] = r(float(bv), 2)

            # EPS for the discrete quarter, not the filed EpsAnnualized column --
            # that one is the cumulative year-to-date figure extrapolated to a
            # full year (BHCL Q2 2024/25: cumulative EPS 5.376 x 4/2 = 10.751,
            # exactly what the column shows), a rate projection rather than what
            # this quarter itself earned. Built instead from the same discrete
            # ProfitAfterTax already on this row, divided by shares outstanding,
            # so it is arithmetically exact against the Net Profit line sitting
            # right above it in the table rather than a second, disagreeing path
            # to a similar-looking number.
            shares = pd.to_numeric(getattr(t, "OrdinaryShares", np.nan), errors="coerce")
            if np.isfinite(shares): last_shares = float(shares)
            pat_q = rec.get("ProfitAfterTax")
            if pat_q is not None and last_shares:
                rec["EPSq"] = r(pat_q * 1000 / last_shares, 2)   # NPR m -> NPR / (shares in '000)
            pat_cum = rec.get("ProfitAfterTax_cum")
            if pat_cum is not None and last_shares:
                rec["EPSq_cum"] = r(pat_cum * 1000 / last_shares, 2)

            # The filed run-rate: year-to-date EPS scaled to a full year by
            # 4/quarter, as filed rather than recomputed -- shown alongside, not
            # instead of, the two above. Naive on this specific fleet: hydro
            # revenue is heavily seasonal within a year, so annualizing a wet
            # quarter overstates and a dry one understates, and the table says so.
            ann = pd.to_numeric(getattr(t, "EpsAnnualized", np.nan), errors="coerce")
            if np.isfinite(ann): rec["EPSann"] = r(float(ann), 2)

            # a null costs six bytes a line across 110 tickers by 8 quarters by
            # 25 lines, and the reader treats absent and null the same way
            qs.append({k: v for k, v in rec.items() if v is not None})
        out[tk] = qs[-8:]
    return out


def _annual(rows, q):
    """One row per fiscal year, full statements rather than a quarter of one.

    Q4's cumulative figure already is the year's total, so there is nothing to
    difference the way a discrete quarter needs -- this reads straight off q,
    the same Audited-preferred, Published-fallback selection _co_payload
    already built for its own register cross-check (Quarter == 4, sorted so
    an audited restatement wins a duplicate over the year's original
    published filing). Reused rather than reselected, so "this year's
    audited figure" means the same row wherever it is read from.
    """
    cols = [c for c, _, _ in CASCADE] + [c for c, _, _ in BS_LINES]
    cols = [c for c in dict.fromkeys(cols) if c in q.columns]
    out = {}
    tickers = {x["tk"] for x in rows}
    for tk, g in q.groupby("Ticker"):
        if tk not in tickers: continue
        yrs = []
        for t in g.itertuples():
            rec = {"y": t.Year, "src": t.DataSource}
            for c in cols:
                cum = pd.to_numeric(getattr(t, c, np.nan), errors="coerce")
                rec[c] = None if not np.isfinite(cum) else r(float(cum)/1e3, 2)
            # Same coalesce _quarters() makes: the filing splits book value
            # across two columns depending on statement type, at most one
            # populated on a given row.
            bv = pd.to_numeric(getattr(t, "BookValuePerShare", np.nan), errors="coerce")
            if not np.isfinite(bv):
                bv = pd.to_numeric(getattr(t, "NetWorthPerShare", np.nan), errors="coerce")
            if np.isfinite(bv): rec["BookValue"] = r(float(bv), 2)
            # Full-year EPS from this row's own Net Profit and share count --
            # not EpsAnnualized, which at Q4 is the cumulative figure times
            # 4/4 and so already equals this to the rounding, but is filed as
            # a run-rate projection, not booked as the year's actual EPS.
            shares = pd.to_numeric(getattr(t, "OrdinaryShares", np.nan), errors="coerce")
            pat = rec.get("ProfitAfterTax")
            if pat is not None and np.isfinite(shares) and shares:
                rec["EPSyr"] = r(pat * 1000 / float(shares), 2)
            yrs.append({k: v for k, v in rec.items() if v is not None})
        out[tk] = yrs[-8:]
    return out


def _quarter_check(quarters, d, cmap):
    """Flag a discrete quarter whose filed energy sales does not look like the
    register's own revenue for the same plants over the same months.

    Found by hand on BHCL's FY2024/25 Q3: 11.4m filed against 33.3m the register
    shows for Upper Machha Khola Small that quarter, while Q1 and Q2 for the same
    company de-cumulate to within 1-2% of the register -- one bad cumulative
    figure in the source, rippling into two displayed quarters (Q4 comes out
    correspondingly overstated). The de-cumulation math itself is not the
    problem; it was checked against those same clean quarters first.

    Reuses _co_payload's own accounting-year re-cut (Shrawan-Asar, not the
    register's Asar-Jestha filing year) rather than a separate one, since that
    re-cut is the one already shown to move the median company from 3.6% off
    its own audited line to 0.6% off it -- inventing a second conversion here
    would be two chances to get the month boundary wrong instead of one.

    A ticker with no plant in the register is left alone: nothing to compare
    reads as unchecked, not as agreeing.

    The register itself runs weeks behind on its most recent months -- this
    build's own newest month has 23 plants filed against a ~147-plant recent
    normal, because most of the fleet simply has not filed it yet. Comparing
    against a quarter built from that would flag nearly every ticker's latest
    quarter at once, in the same direction, which is filing lag wearing the
    costume of a data error and would bury the one real signal (BHCL's, an
    isolated quarter surrounded by clean ones) under ~85 false ones. A quarter
    is only compared once every one of its three months has close to the
    fleet's typical recent filing count -- not "some data", enough of it.
    """
    pids = {t.Ticker: [int(x) for x in str(t.PlantIds).split(";") if x]
            for t in cmap.itertuples() if t.PlantIds}
    mo = d[~d.IsAnnualFiling].copy()
    mo["acct"] = np.where(mo.BsMonth >= 4, mo.BsYear, mo.BsYear - 1)
    mo["amonth"] = (mo.BsMonth - 4) % 12 + 1              # 1 = Shrawan, 12 = Asar
    mo["aq"] = (mo.amonth - 1) // 3 + 1                    # NEPSE Q1-Q4 on that year
    # AD "2024/25" is BS 2081/82 throughout this codebase (_co_payload's own
    # figure); the AD start-year is the BS accounting year minus 57.
    mo["fy0"] = mo.acct - 57

    # A high percentile over the trailing window, not the median: a couple of
    # still-filing months sitting in that same window should not drag down the
    # bar a month has to clear to be judged complete.
    monthly_n = d.groupby(["BsYear", "BsMonth"]).PlantName.nunique()
    typical = monthly_n.tail(18).quantile(0.75)
    floor = 0.6 * typical
    complete_bsm = {ym for ym, n in monthly_n.items() if n >= floor}

    def quarter_complete(fy0, aq):
        bs = fy0 + 57
        months = {"Q1": [(bs,4),(bs,5),(bs,6)], "Q2": [(bs,7),(bs,8),(bs,9)],
                  "Q3": [(bs,10),(bs,11),(bs,12)], "Q4": [(bs+1,1),(bs+1,2),(bs+1,3)]
                  }["Q"+str(aq)]
        return all(ym in complete_bsm for ym in months)

    for tk, qs in quarters.items():
        plant_ids = pids.get(tk)
        if not plant_ids: continue
        rms = (mo[mo.PlantId.isin(plant_ids)]
                 .groupby(["fy0", "aq"]).Revenue_NPR.sum() / 1e6)   # NPR million
        for q in qs:
            es = q.get("EnergySales")
            if es is None: continue
            fy0, aq = int(str(q["y"]).split("/")[0]), q["q"]
            if not quarter_complete(fy0, aq): continue
            got = rms.get((fy0, aq))
            if got is None or got <= 1: continue    # nothing worth comparing to
            ratio = es / got
            q["rms_es"] = r(float(got), 2)
            # Wide on purpose: this is "something is clearly wrong", not a
            # disagreement-with-the-register meter -- that is what section 06's
            # reconciliation and the "agree" histogram above are already for.
            # A single quarter is noisier than an annual figure, and real
            # accrual-timing gaps exist without either source being wrong.
            if ratio < 0.5 or ratio > 2.0:
                q["es_suspect"] = 1


def _growth(rows, quarters):
    """Year to date against the prior year's Q4, and this quarter against the same
    quarter a year ago -- which is the comparison a cumulative filing supports
    without its own seasonality getting in the way."""
    try:
        f = pd.read_csv(P("scalper_hydro_combined.csv"))
    except FileNotFoundError:
        return
    f["_aud"] = (f.DataSource == "Audited").astype(int)
    f = f.sort_values("_aud").drop_duplicates(["Ticker","Year","Quarter"], keep="last")
    f["fy0"] = f.Year.map(lambda s: int(str(s).split("/")[0]))
    by = {x["tk"]: x for x in rows}
    for tk, g in f.groupby("Ticker"):
        x = by.get(tk)
        if x is None: continue
        g = g.sort_values(["fy0","Quarter"])
        last = g.iloc[-1]
        x["lastq"] = f"{last.Year} Q{int(last.Quarter)}"
        for tag, col in (("rev","EnergySales"), ("pat","ProfitAfterTax")):
            cur = pd.to_numeric(last[col], errors="coerce")
            if not np.isfinite(cur) or cur == 0: continue
            yoy = g[(g.fy0 == last.fy0 - 1) & (g.Quarter == last.Quarter)]
            if len(yoy):
                p = pd.to_numeric(yoy.iloc[0][col], errors="coerce")
                if np.isfinite(p) and p > 0: x[f"yoy_{tag}"] = r(cur/p - 1, 4)
            q4 = g[(g.fy0 == last.fy0 - 1) & (g.Quarter == 4)]
            if len(q4) and last.Quarter < 4:
                p = pd.to_numeric(q4.iloc[0][col], errors="coerce")
                if np.isfinite(p) and p > 0: x[f"ytd_{tag}"] = r(cur/p - 1, 4)


def _dividends(tickers):
    """Declared dividends per ticker, newest first, plus the sector by year.

    nepse_dividends.csv is every listed company on the exchange; only the hydro
    tickers are read out of it. Percentages are of par, and par is NPR 100
    throughout NEPSE, so a cash figure of 0.53% is 53 paisa a share.

    Bonus and cash are kept apart and never added into one number on the page.
    They are not the same transaction: cash leaves the company, a bonus issue
    prints shares against reserves and leaves the balance sheet where it was.
    Reading a 10% bonus as a 10% dividend is the easiest mistake to make with
    this file, and the sector has spent a decade moving from one to the other --
    a total that hid the shift would hide the finding.
    """
    try:
        dv = pd.read_csv(P("nepse_dividends.csv"), dtype=str).fillna("")
    except FileNotFoundError:
        return {}, []
    dv = dv.rename(columns={"Bonus (%)": "bonus", "Cash (%)": "cash",
                            "Announcement Date": "ann", "Fiscal Year": "fy"})
    dv = dv[dv.Symbol.isin(set(tickers))].copy()
    if dv.empty: return {}, []
    # Unilever files its cash with a thousands separator; nothing else in the
    # file does, and nothing hydro does, but the parse has to survive it anyway
    for c in ("bonus", "cash"):
        dv[c] = pd.to_numeric(dv[c].str.replace(",", ""), errors="coerce").fillna(0.0)
    # "2081/2082" -> BS 2081, the year the financials call AD 2024/25
    dv["y"] = pd.to_numeric(dv.fy.str.slice(0, 4), errors="coerce")
    dv = dv.dropna(subset=["y"]).sort_values("y", ascending=False)

    per = {tk: [[int(t.y), r(t.bonus, 2), r(t.cash, 2), t.ann or None]
                for t in g.itertuples()]
           for tk, g in dv.groupby("Symbol")}

    by_year = []
    for y, g in dv.groupby(dv.y.astype(int)):
        by_year.append({"y": int(y), "n": int(g.Symbol.nunique()),
                        "cash": r(g.cash.median(), 3), "bonus": r(g.bonus.median(), 3),
                        # a cash line under one per cent of par is the withholding
                        # on the bonus issue beside it, not a distribution
                        "n_cash": int((g.cash > 1).sum()),
                        "n_bonus": int((g.bonus > 0).sum())})
    return per, by_year


def _co_payload(d, m):
    """Every listed hydro ticker, with the register's meter beside its own books.

    Three files meet here and no two of them share a key. The financials are
    filed per NEPSE ticker and per AD fiscal year; the register is filed per
    plant and per BS fiscal year; the load factors are hand-entered per ticker.
    scalper_company_map.csv is the bridge, and the two calendars are the same
    year offset by 57 -- AD 2024/25 is BS 2081/82.

    The one thing that has to be right is which twelve months to compare. A
    register fiscal year runs Asar to Jestha; a Nepali accounting year runs
    Shrawan to Asar, one month later. Re-cutting the monthly rows onto the
    accounting year moves the median company from 3.6% off its own audited
    energy-sales line to 0.6% off it, and takes the count agreeing within two
    per cent from 102 to 192 of 310. So the re-cut is not a refinement; it is
    the difference between the two sources agreeing and not.

    Every ticker comes back, including the 8 with no plant in the register and
    the ones whose only filed year is still running. A company that has not
    generated yet is a fact about the sector, not a row to drop.
    """
    try:
        fin  = pd.read_csv(P("scalper_hydro_combined.csv"))
        cmap = pd.read_csv(P("scalper_company_map.csv")).fillna("")
    except FileNotFoundError:
        return None
    try:
        mkt = pd.read_csv(P("scalper_hydro_comparative.csv")).set_index("Ticker")
    except FileNotFoundError:
        mkt = None
    try:
        pin = pd.read_csv(P("scalper_plf_input.csv"), comment="#",
                          dtype=str).fillna("").set_index("Ticker")
    except FileNotFoundError:
        pin = None

    # Q4 is the year: every line in this series is cumulative year to date.
    # Where a year was filed twice the audited figure is the one that stands --
    # 33 of 268 restatements move the top line by more than a per cent.
    q = fin[fin.Quarter == 4].copy()
    q["_aud"] = (q.DataSource == "Audited").astype(int)
    q = q.sort_values("_aud").drop_duplicates(["Ticker", "Year"], keep="last")
    # and in year order, because the latest complete year is read off the end
    q = q.sort_values("Year", key=lambda c: c.map(lambda y: int(str(y).split("/")[0])))

    pids = {t.Ticker: [int(x) for x in str(t.PlantIds).split(";") if x]
            for t in cmap.itertuples() if t.PlantIds}
    cm = cmap.set_index("Ticker")

    mo = d[~d.IsAnnualFiling].copy()
    mo["acct"] = np.where(mo.BsMonth >= 4, mo.BsYear, mo.BsYear - 1)
    ay = (mo.groupby(["PlantId", "acct"], as_index=False)
            .agg(gen=("Generation_kWh","sum"), rev=("Revenue_NPR","sum"),
                 roy=("Royalty_NPR","sum"), n=("Period","size")))
    # the register's own cut, kept only to show what the re-cut is worth
    rf = (mo.groupby(["PlantId", "FiscalYear"], as_index=False)
            .agg(rev=("Revenue_NPR","sum"), n=("Period","size")))
    kw = d.groupby("PlantId").Capacity_kW.first()

    def g(row, col, scale=1.0, nd=2):
        v = pd.to_numeric(row.get(col), errors="coerce") if hasattr(row, "get") else None
        return None if v is None or not np.isfinite(v) else r(v*scale, nd)

    hist, fleet, ratios = {}, {}, []
    for t in q.itertuples():
        ids = pids.get(t.Ticker)
        if not ids: continue
        bs = int(str(t.Year).split("/")[0]) + 57
        fy = f"{bs-2000:03d}/{(bs+1)-2000:02d}"
        A = ay[(ay.PlantId.isin(ids)) & (ay.acct == bs)]
        if len(A) != len(ids) or not (A.n >= 12).all() or A.rev.sum() <= 0: continue
        cap = float(kw[ids].sum())
        rep = float(t.EnergySales)/1e3 if np.isfinite(t.EnergySales) and t.EnergySales > 0 else None
        reg = float(A.rev.sum())/1e6
        gwh = float(A.gen.sum())/1e6
        plf = A.gen.sum()/(cap*8760) if cap > 0 else None
        hist.setdefault(t.Ticker, []).append(
            [t.Year, r(reg,1), r(rep,1), r(gwh,1), r(plf,3),
             g(t._asdict(), "ProfitAfterTax", 1e-3, 1), r(A.roy.sum()/1e6, 2)])
        if rep:
            f = fleet.setdefault(t.Year, {"ad": t.Year, "fy": fy, "reg": 0.0, "rep": 0.0, "n": 0})
            f["reg"] += reg; f["rep"] += rep; f["n"] += 1
            F = rf[(rf.PlantId.isin(ids)) & (rf.FiscalYear == fy)]
            ratios.append({"tk": t.Ticker, "ad": t.Year, "a": reg/rep,
                           "f": (F.rev.sum()/1e6/rep) if len(F) == len(ids)
                                 and (F.n >= 11).all() and F.rev.sum() > 0 else None})

    # ── one row per ticker, latest complete year, everything alongside
    divs, div_year = _dividends(set(q.Ticker))
    rows = []
    for tk in sorted(set(q.Ticker)):
        h = hist.get(tk, [])
        last = h[-1] if h else None
        cr = cm.loc[tk] if tk in cm.index else None
        mr = mkt.loc[tk] if mkt is not None and tk in mkt.index else None
        pr = pin.loc[tk] if pin is not None and tk in pin.index else None
        ids = pids.get(tk, [])
        mw = float(kw[[i for i in ids if i in kw.index]].sum())/1000 if ids else None
        if not mw and cr is not None:
            mw = pd.to_numeric(cr.get("MW"), errors="coerce")
            mw = None if not np.isfinite(mw) else float(mw)
        cplf = float(pr.ContractPLF)/100 if pr is not None and pr.ContractPLF else None
        dplf = float(pr.DesignPLF)/100 if pr is not None and pr.DesignPLF else None
        mcap = None if mr is None else g(mr, "Market Cap", 1e-6, 0)     # NPR -> million
        rec = {
            "tk": tk,
            "name": (str(cr.Company) if cr is not None else tk)[:46],
            "reg_co": (str(cr.RegisterCompany) or None) if cr is not None else None,
            "mw": r(mw, 2), "np": len(ids),
            "pnames": (str(cr.PlantNames) or None) if cr is not None else None,
            "dist": (str(cr.District) or None) if cr is not None else None,
            # register side, latest complete accounting year
            "fy": last[0] if last else None,
            "reg_m": last[1] if last else None, "rep_m": last[2] if last else None,
            "gwh": last[3] if last else None, "plf": last[4] if last else None,
            "roy_m": last[6] if last else None,
            "rpm": r(last[1]/mw, 1) if last and last[1] and mw else None,
            "ratio": r(last[1]/last[2], 3) if last and last[1] and last[2] else None,
            "nyr": len(h),
            # what the plant was built or contracted to do
            "cplf": r(cplf, 4), "dplf": r(dplf, 4),
            "cost_mw": r(float(pr.CostPerMW), 1) if pr is not None and pr.CostPerMW else None,
            "psrc": (pr.Source or None) if pr is not None else None,
            # the company's own books and what the market pays for them
            "pat_m": last[5] if last else None,
            "mcap_m": mcap, "mcap_mw": r(mcap/mw, 0) if mcap and mw else None,
            "roe": None if mr is None else g(mr, "ROE (TTM) %", 1, 4),
            "pe": None if mr is None else g(mr, "PE (TTM)", 1, 1),
            "pbv": None if mr is None else g(mr, "PBV", 1, 2),
            "npm": None if mr is None else g(mr, "Net Profit Margin TTM %", 1, 4),
            "d2a": None if mr is None else g(mr, "Debt to Asset", 1, 3),
            "px": None if mr is None else g(mr, "Latest Close", 1, 1),
            "yr1": None if mr is None else g(mr, "1 Year Price Chg %", 1, 3),
            "hist": h,
            # newest first: [BS year, bonus % of par, cash % of par, announced]
            "div": divs.get(tk, []),
        }
        dl = rec["div"][0] if rec["div"] else None
        rec |= {"div_y": dl[0] if dl else None,
                "bonus": dl[1] if dl else None, "cash": dl[2] if dl else None,
                "div_n": len(rec["div"]),
                # cash on par 100 against the traded price; a bonus issue is not
                # a yield and is deliberately left out of it
                "yld": r(dl[2]/rec["px"], 4) if dl and dl[2] and rec.get("px") else None}
        rec["ach"] = r(rec["plf"]/(cplf or dplf), 3) if rec["plf"] and (cplf or dplf) else None
        rec["st"] = "ok" if last else ("part" if ids else "off")
        rows.append(rec)

    if not any(x["st"] == "ok" for x in rows): return None

    # ── how close the two sources land, on each cut of the year
    A = [x["a"] for x in ratios]
    both = [(x["a"], x["f"]) for x in ratios if x["f"] is not None]
    within = lambda v, e: sum(1 for x in v if abs(x-1) <= e)
    agree = {
        "n": len(A), "n_both": len(both), "tickers": len({x["tk"] for x in ratios}),
        "med": r(float(np.median(A)), 4),
        "p25": r(float(np.percentile(A, 25)), 3), "p75": r(float(np.percentile(A, 75)), 3),
        "w2": within(A, .02), "w5": within(A, .05), "w10": within(A, .10),
        # the same company-years on the register's own cut, so the pair is fair
        "med_a": r(float(np.median([abs(np.log(a)) for a, f in both])), 4),
        "med_f": r(float(np.median([abs(np.log(f)) for a, f in both])), 4),
        "w2_a": within([a for a, f in both], .02), "w2_f": within([f for a, f in both], .02),
    }
    mkt = _market(rows)
    quarters = _quarters(rows)
    _quarter_check(quarters, d, cmap)
    _growth(rows, quarters)
    annual = _annual(rows, q)

    order = sorted(fleet, key=lambda y: int(str(y).split("/")[0]))
    paid = [x for x in rows if x["div_n"]]
    return {"rows": rows, "agree": agree, "divYear": div_year,
            "mkt": mkt, "q": quarters, "annual": annual,
            "cascade": [[c, s, lab] for c, s, lab in CASCADE],
            "bs": [[c, lab, sub] for c, lab, sub in BS_LINES],
            "n_div": len(paid),
            "n_div_now": sum(1 for x in rows if x["div_y"] == (div_year[-1]["y"] if div_year else None)),
            "fleet": [{"ad": y, "fy": fleet[y]["fy"], "reg_bn": r(fleet[y]["reg"]/1e3, 2),
                       "rep_bn": r(fleet[y]["rep"]/1e3, 2), "n": fleet[y]["n"]} for y in order],
            "n_ok": sum(1 for x in rows if x["st"] == "ok"),
            "n_part": sum(1 for x in rows if x["st"] == "part"),
            "n_off": sum(1 for x in rows if x["st"] == "off"),
            "n_plf": sum(1 for x in rows if x["ach"] is not None),
            "mw_ok": r(sum(x["mw"] or 0 for x in rows if x["st"] == "ok"), 0),
            "mw_all": r(sum(x["mw"] or 0 for x in rows), 0)}


def _recon_payload(d, m):
    """Tie RMS billing for Likhu-4 to its operator's reported energy revenue.
    Two adjustments make it line up: RMS files each block Asar-Jestha, one month
    ahead of the Shrawan-Asar accounting year, and the published quarterlies are
    cumulative year-to-date."""
    cum = _reported_quarters()
    if not cum: return None
    p = d[d.PlantId == 116].copy(); p["idx"] = p.BsYear*12 + p.BsMonth
    rev = p.groupby("idx").Revenue_NPR.sum()
    QN = ["Q1","Q2","Q3","Q4"]
    QM = {"Q1":[4,5,6], "Q2":[7,8,9], "Q3":[10,11,12], "Q4":[1,2,3]}
    rows = []
    # Derived, not listed. The listed version was written for three years and then
    # silently ignored the two the cache had already grown by, dropping seven
    # quarters that reconcile -- three of them to within 0.005%.
    for ad in sorted(cum):
        bs = 2080 + (int(str(ad).split("/")[0]) - 2023)
        for i, q in enumerate(QN):
            if q not in cum.get(ad, {}): continue
            standalone = cum[ad][q] - (cum[ad].get(QN[i-1]) or 0 if i else 0)
            got = sum(rev.get((bs if mm >= 4 else bs+1)*12 + mm, 0) for mm in QM[q])
            if got == 0 or not standalone: continue
            rows.append({"label": f"{ad} {q}", "rms": round(got), "rep": round(standalone),
                         "diff": round(got-standalone),
                         "pct": round(100*(got-standalone)/standalone, 2)})
    mr = m[m.PlantId == 116]
    return {"plant": "Likhu-4", "mw": 52.4, "rows": rows,
            "company": None if mr.empty else str(mr.CompanyId.iloc[0]),
            "cod": None if mr.empty else str(mr.MitiofOperation.iloc[0]),
            "n_quarters": len(rows),
            "n_exact": sum(1 for q in rows if abs(q["pct"]) < 0.05),
            "median_abs_pct": round(float(np.median([abs(q["pct"]) for q in rows])), 3)
                              if rows else None}



# ═════════════════════════════════════════════════════ standalone map export ══
def export_map_svg(path="nepal_fleet_map.svg", light=True):
    """Render the fleet map to a self-contained SVG, same projection as the page.
    Useful for sharing or printing without the dashboard around it."""
    pl = json.load(open(P("dashboard.json"), encoding="utf-8"))
    M = pl["map"]
    W, H, pad = 1200, 700, 16
    C = (dict(bg="#ffffff", land="#f4f7f8", line="#d3dde1", river="#8fb8d0",
              peak="#a89f92", pink="#6d6459", s1="#2a78d6", s2="#eb6834", ink="#5f7883",
              grid="#8b5fb0")
         if light else
         dict(bg="#0b1317", land="#17252c", line="#26383f", river="#37596e",
              peak="#5a5348", pink="#8c8477", s1="#3987e5", s2="#d95926", ink="#7b939d",
              grid="#a273c9"))
    lon0, lat0, lon1, lat1 = M["bbox"]
    k  = math.cos(math.radians(28.4))
    sc = min((W-2*pad)/((lon1-lon0)*k), (H-2*pad)/(lat1-lat0))
    ox = (W - (lon1-lon0)*k*sc)/2
    oy = (H - (lat1-lat0)*sc)/2
    X  = lambda lo: ox + (lo-lon0)*k*sc
    Y  = lambda la: oy + (lat1-la)*sc
    out = [f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {W} {H}" width="{W}" '
           f'height="{H}" font-family="IBM Plex Mono, ui-monospace, monospace">',
           f'<rect width="{W}" height="{H}" fill="{C["bg"]}"/>']
    for ring in M["outline"]:
        d = "".join(("L" if i else "M")+f"{X(p[0]):.1f} {Y(p[1]):.1f}" for i, p in enumerate(ring))
        out.append(f'<path d="{d}Z" fill="{C["land"]}" stroke="{C["line"]}" stroke-width="1.4"/>')
    for rv in M.get("rivers", []):
        w = 1.9 if rv["km"] >= 200 else (1.4 if rv["km"] >= 100 else (1.0 if rv["km"] >= 50 else .7))
        for seg in rv["segs"]:
            d = "".join(("L" if i else "M")+f"{X(p[0]):.1f} {Y(p[1]):.1f}" for i, p in enumerate(seg))
            out.append(f'<path d="{d}" fill="none" stroke="{C["river"]}" stroke-width="{w}" '
                       f'stroke-linecap="round" stroke-linejoin="round" opacity=".85"/>')
    # RPGCL's own corridors, drawn first so the OSM-built network sits on top.
    # These arrive as dash segments and are left that way - planned should look
    # planned. Every class here is read off the sheet's own ArcMap layer tag
    # (see extract_rpgcl.py), not guessed from colour, which is the only way
    # its 132 kV existing/under-construction/proposed corridors can be told
    # apart at all - the source PDF draws all three in the same red.
    PLAN_STY = {
        "existing_400":     ("#9a4fc4", 1.9, .85), "cross_border_400": ("#e0507a", 1.8, .85),
        "proposed_400":     ("#c765d9", 1.4, .65), "existing_220":     ("#3f6fc9", 1.5, .85),
        "uc_220":           ("#5a93d6", 1.3, .72), "proposed_220":     ("#6d8fd6", 1.05, .6),
        "existing_132":     ("#c9822e", 1.25, .85), "uc_132":          ("#d99a4e", 1.1, .7),
        "proposed_132":     ("#e8b073", 0.85, .5),
    }
    for ln in M.get("plan", []):
        cc, w, op = PLAN_STY.get(ln["c"], ("#b07cd8", 1.3, .7))
        d = "".join(("L" if i else "M")+f"{X(p[0]):.1f} {Y(p[1]):.1f}"
                    for i, p in enumerate(ln["pts"]))
        out.append(f'<path d="{d}" fill="none" stroke="{cc}" stroke-width="{w}" '
                   f'stroke-linecap="round" opacity="{op}"/>')

    # transmission network: violet keeps it clear of the blue rivers, the grey
    # peaks and the blue/orange plant markers
    for ln in M.get("lines", []):
        kv = ln["kv"]
        w  = 2.2 if kv >= 400 else (1.6 if kv >= 220 else (1.0 if kv >= 132 else 0.6))
        op = .9  if kv >= 400 else (.8  if kv >= 220 else (.62 if kv >= 132 else .4))
        d = "".join(("L" if i else "M")+f"{X(p[0]):.1f} {Y(p[1]):.1f}"
                    for i, p in enumerate(ln["pts"]))
        out.append(f'<path d="{d}" fill="none" stroke="{C["grid"]}" stroke-width="{w}" '
                   f'stroke-linecap="round" stroke-linejoin="round" opacity="{op}"/>')
    for sb in M.get("subs", []):
        x, y = X(sb["lo"]), Y(sb["la"])
        out.append(f'<rect x="{x-2.6:.1f}" y="{y-2.6:.1f}" width="5.2" height="5.2" '
                   f'fill="none" stroke="{C["grid"]}" stroke-width="1.1" opacity=".75"/>')
    # pipeline projects: diamonds, so they never read as built plants (circles)
    # or substations (squares). Sized by capacity where the sheet prints one.
    TIER_C = {"built": "#2f9e44", "licensed": "#74c476",
              "survey": "#6fa8dc", "applied": "#c39bd3"}
    for pp in M.get("plan_plants", []):
        x, y = X(pp["lo"]), Y(pp["la"])
        r0 = 2.4 + (0.55 * math.sqrt(pp["mw"]) if pp.get("mw") else 0)
        r0 = min(r0, 13)
        cc = TIER_C.get(pp.get("tier"), "#999")
        out.append(f'<path d="M{x:.1f} {y-r0:.1f}L{x+r0:.1f} {y:.1f}'
                   f'L{x:.1f} {y+r0:.1f}L{x-r0:.1f} {y:.1f}Z" fill="{cc}" '
                   f'fill-opacity=".55" stroke="{cc}" stroke-width="1"/>')

    # small-hydro clusters: a downward triangle, so it is never confusable with
    # the pipeline diamonds, the plant circles or the peaks' upward triangles.
    for cl in M.get("plan_clusters", []):
        x, y = X(cl["lo"]), Y(cl["la"])
        r0 = min(11, 2.8 + (0.46 * math.sqrt(cl["mw"]) if cl.get("mw") else 0))
        out.append(f'<path d="M{x-r0:.1f} {y-r0*.7:.1f}L{x+r0:.1f} {y-r0*.7:.1f}'
                   f'L{x:.1f} {y+r0*.85:.1f}Z" fill="#0d7bd9" fill-opacity=".45" '
                   f'stroke="#0d7bd9" stroke-width="1"/>')

    # RPGCL substations, filled squares coloured by build status as on the sheet
    SUB_C = {"existing": "#d94a4a", "under_construction": "#27b34a", "future": "#3b7fe0"}
    for sb in M.get("plan_subs", []):
        x, y = X(sb["lo"]), Y(sb["la"])
        cc = SUB_C.get(sb["c"], "#888")
        out.append(f'<rect x="{x-3:.1f}" y="{y-3:.1f}" width="6" height="6" fill="{cc}" '
                   f'fill-opacity=".9" stroke="{C["bg"]}" stroke-width="0.8"/>')

    placed = []
    for pk in M.get("peaks", []):
        x, y = X(pk["lo"]), Y(pk["la"])
        sz = 8 if pk["e"] >= 8000 else (6 if pk["e"] >= 7500 else 4.5)
        op = .95 if pk["e"] >= 8000 else .6
        out.append(f'<path d="M{x:.1f} {y-sz:.1f}L{x+sz*.92:.1f} {y+sz*.72:.1f}'
                   f'L{x-sz*.92:.1f} {y+sz*.72:.1f}Z" fill="{C["peak"]}" opacity="{op}"/>')
        if pk["e"] < 8000: continue
        nm = pk["n"].replace("Mount ", "")
        lw, lx, ly = len(nm)*6.2+8, x, y-sz-6
        if any(abs(q[0]-lx) < (q[2]+lw)/2 and abs(q[1]-ly) < 13 for q in placed):
            ly = y+sz+14
        if any(abs(q[0]-lx) < (q[2]+lw)/2 and abs(q[1]-ly) < 13 for q in placed): continue
        placed.append((lx, ly, lw))
        out.append(f'<text x="{lx:.1f}" y="{ly:.1f}" text-anchor="middle" fill="{C["pink"]}" '
                   f'font-size="11" font-weight="600">{nm}</text>')
    for p in sorted(M["pts"], key=lambda q: -(q["mw"] or 0)):
        x, y = X(p["lo"]), Y(p["la"])
        rr = min(20, 2.6 + .95*math.sqrt(p["mw"] or .5))
        col = C["s2"] if p["t2"] else C["s1"]
        if p["p"] == 2:
            out.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="{rr:.1f}" fill="{col}" '
                       f'fill-opacity=".85" stroke="{C["bg"]}" stroke-width="1.6"/>')
        elif p["p"] == 1:
            out.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="{rr:.1f}" fill="{col}" '
                       f'fill-opacity=".38" stroke="{col}" stroke-width="1.4"/>')
        else:
            out.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="{rr:.1f}" fill="none" '
                       f'stroke="{col}" stroke-width="1.4" stroke-opacity=".6" '
                       f'stroke-dasharray="3 2.4"/>')
    out.append(f'<text x="{pad+8}" y="{H-16}" fill="{C["ink"]}" font-size="12">'
               f'{len(M["pts"])} plants · {len(M.get("rivers",[]))} rivers · '
               f'{len(M.get("peaks",[]))} peaks · {len(M.get("lines",[]))} grid lines '
               f'· DoED RMS + OpenStreetMap</text>')
    out.append("</svg>")
    svg = "\n".join(out)
    open(P(path), "w", encoding="utf-8").write(svg)
    print(f"wrote {path} ({len(svg.encode('utf-8')):,} bytes)")
    return path



# ══════════════════════════════════════════════════ 3D terrain map (local) ══
def export_map3d(path="nepal_fleet_3d.html", tpl="map3d.tpl.html"):
    """MapLibre on real elevation tiles, written from map3d.tpl.html.

    This CANNOT be published as an Artifact: the Artifact CSP blocks every
    external host, so the tile and script requests fail and the map renders
    blank. It is a local file - open it directly, or serve it over localhost.
    Tile sources are keyless (OpenStreetMap raster + Mapzen/AWS terrarium DEM);
    respect their usage policies if you point many people at it.
    """
    pl = json.load(open(P("dashboard.json"), encoding="utf-8"))
    M = pl["map"]
    keep = ("n", "co", "mw", "la", "lo", "p", "src", "di", "cod", "age", "t2", "gwh")
    pts = [{k: p.get(k) for k in keep} for p in M["pts"]]
    rivers = M.get("rivers", [])
    peaks  = M.get("peaks", [])
    lines  = M.get("lines", [])
    subs   = M.get("subs", [])
    plan   = M.get("plan", [])
    psubs  = M.get("plan_subs", [])
    pplants= M.get("plan_plants", [])
    pclus  = M.get("plan_clusters", [])
    J = lambda o: json.dumps(o, separators=(",", ":"))
    html = (open(P(tpl), encoding="utf-8").read()
            .replace("__PLANTS__", J(pts))
            .replace("__RIVERS__", J(rivers))
            .replace("__PEAKS__", J(peaks))
            .replace("__LINES__", J(lines))
            .replace("__SUBS__", J(subs))
            .replace("__PLAN__", J(plan))
            .replace("__PSUBS__", J(psubs))
            .replace("__PPLANTS__", J(pplants))
            .replace("__PCLUS__", J(pclus))
            .replace("__NPP__", str(len(pplants)))
            .replace("__NPC__", str(len(pclus)))
            .replace("__NPS__", str(len(psubs)))
            .replace("__NL__", str(len(lines)))
            .replace("__N__", str(len(pts)))
            .replace("__NR__", str(len(rivers)))
            .replace("__NP__", str(len(peaks))))
    open(P(path), "w", encoding="utf-8").write(html)
    print(f"wrote {path} ({len(html.encode('utf-8')):,} bytes, {len(pts)} plants,"
          f" {len(rivers)} rivers, {len(peaks)} peaks, {len(lines)} grid lines)")
    print("  local file only - Artifact CSP blocks its tile requests")
    return path


# ═══════════════════════════════════════════════════════════════════ render ══
def cmd_build(write=True):
    pl = build_payload()
    st, rg = pl["stats"], pl["regime"]
    print(f"plants {st['plants']} | plant-years {st['plant_years']} | months {st['month_rows']}"
          f" | fy {st['fy_span']}")
    print(f"capacity {st['mw']} MW | generation {st['twh']} TWh | royalty {st['royalty_bn']} bn")
    print(f"commissioning dates parsed {st['cod_known']}/{st['cod_total']}")
    print(f"transitions {rg['n_trans']} (exactly 15 yrs: {rg['n_exact15']}) | "
          f"under 15: {rg['under15_high_pct']}% | 15+: {rg['over15_high_pct']}%")
    if "map" in pl:
        mp = pl["map"]
        print(f"map {len(mp['pts'])} plotted, {mp['linked']} linked, {mp['n_exact']} exact /"
              f" {mp['n_district']} district / {mp['n_none']} unlocated")
    if "recon" in pl:
        rc = pl["recon"]
        print(f"recon {rc['n_quarters']} quarters, {rc['n_exact']} within 0.05%,"
              f" median |diff| {rc['median_abs_pct']}%")
    if not write: return pl
    tpl = open(P("template.html"), encoding="utf-8").read()
    if "__PAYLOAD__" not in tpl:
        raise SystemExit("template.html has no __PAYLOAD__ placeholder")
    html = tpl.replace("__PAYLOAD__", json.dumps(pl, separators=(",", ":"), ensure_ascii=False))
    open(P("nepal_hydro.html"), "w", encoding="utf-8").write(html)
    json.dump(pl, open(P("dashboard.json"), "w", encoding="utf-8"),
              separators=(",", ":"), ensure_ascii=False)
    print(f"\nwrote nepal_hydro.html ({len(html.encode('utf-8')):,} bytes)")
    return pl


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    cmd = args[0] if args else "build"
    if cmd == "scrape": cmd_scrape()
    elif cmd == "latest": cmd_scrape(latest_only=True)
    elif cmd == "clean": clean()
    elif cmd == "meta": cmd_meta()
    elif cmd == "geo": cmd_geo()
    elif cmd == "coords": resolve_coords()
    elif cmd == "companies": cmd_companies()
    elif cmd == "terrain": fetch_terrain()
    elif cmd == "grid": fetch_grid()
    elif cmd == "mapsvg": export_map_svg()
    elif cmd == "map3d": export_map3d()
    elif cmd == "all":
        cmd_scrape(); cmd_meta(); cmd_geo(); cmd_build()
    elif cmd == "build": cmd_build(write="--stats" not in sys.argv)
    else: raise SystemExit(__doc__)
